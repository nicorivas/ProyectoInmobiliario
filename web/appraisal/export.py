
from django.http import HttpResponse
from django.core import files

import requests
import tempfile

from realestate.models import RealEstate, Asset
from house.models import House
#from building.models import Building
from apartment.models import Apartment
from appraisal.models import Appraisal, Comment, Photo, Document
from commune.models import Commune
from user.models import UserProfile

import os

from openpyxl import drawing
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook

from django.db.models import ManyToOneRel, ManyToManyRel

import datetime

def export(request,forms,appraisal,realEstate):

    module_dir = os.path.dirname(__file__)  # get current directory
    file_path = os.path.join(module_dir,'static/appraisal/santander-template.xlsx')

    def excel_find(workbook,term):
        for sheet in workbook:
            for row in range(1,400):
                for col in range(1,100):
                   cv = sheet.cell(row=row,column=col).value

    def excel_find_replace(workbook,term,rep,once=False):
        for sheet in workbook:
            for row in range(1,400):
                for col in range(1,100):
                   cv = sheet.cell(row=row,column=col).value
                   if cv == term:
                       sheet.cell(row=row,column=col).value = rep
                       if once: return

    wb = load_workbook(filename=file_path)

    fields = Appraisal._meta.get_fields()
    for field in fields:
        if not isinstance(field,ManyToOneRel):
            field_name = field.deconstruct()[0]
            if field_name == 'tasadorUser':
                if isinstance(appraisal.tasadorUser,type(None)):
                    excel_find_replace(wb,field_name,'')
                else:
                    userProfile = UserProfile.objects.get(user=appraisal.tasadorUser)
                    excel_find_replace(wb,field_name,userProfile.full_name)
                    excel_find_replace(wb,field_name+'.rut',userProfile.rut_verbose)
                continue
            if len(field.choices) == 0:
                excel_find_replace(wb,field_name,getattr(appraisal,field_name))
            else:
                value = getattr(appraisal,field_name)
                for a in field.choices:
                    if a[0] == value:
                        excel_find_replace(wb,field_name,a[1])
                        break

    fields = RealEstate._meta.get_fields()
    for field in fields:
        if not isinstance(field,ManyToOneRel) and not isinstance(field,ManyToManyRel):
            field_name = field.deconstruct()[0]
            if field_name == 'tasadorUser':
                continue
            if field_name == 'antiguedad':
                continue
            if field_name == 'marketPrice':
                continue
            if field_name == 'addressCommune':
                excel_find_replace(wb,'addressCommune',realEstate.addressCommune.name)
            if field_name == 'addressRegion':
                excel_find_replace(wb,'addressRegion',realEstate.addressRegion.name)
            if len(field.choices) == 0:
                excel_find_replace(wb,field_name,getattr(realEstate,field_name))
            else:
                value = getattr(realEstate,field_name)
                for a in field.choices:
                    if a[0] == value:
                        excel_find_replace(wb,field_name,a[1])
                        break

    year = datetime.datetime.now().year
    antiguedad = year-realEstate.anoConstruccion
    excel_find_replace(wb,'antiguedad','± '+str(antiguedad)+' años')
    excel_find_replace(wb,'vidaUtil','± '+str(realEstate.vidaUtil)+' años')

    excel_find_replace(wb,'address',realEstate.house.addressVerboseNoRegionNoCommune)

    if not isinstance(realEstate.permisoEdificacionFecha,type(None)):
        excel_find_replace(wb,'permisoEdificacion',realEstate.permisoEdificacionNo+' del año '+str(realEstate.permisoEdificacionFecha.year))
    else:
        excel_find_replace(wb,'permisoEdificacion',realEstate.permisoEdificacionNo)
    if not isinstance(realEstate.recepcionFinalFecha,type(None)):
        excel_find_replace(wb,'recepcionFinal',realEstate.recepcionFinalNo+' del año '+str(realEstate.recepcionFinalFecha.year))
    else:
        excel_find_replace(wb,'recepcionFinal',realEstate.recepcionFinalNo)

    if realEstate.is_house:
        excel_find_replace(wb,'generalDescription',realEstate.house.generalDescription)
    else:
        excel_find_replace(wb,'generalDescription',realEstate.apartment.generalDescription)

    excel_find_replace(wb,'descripcionSectorAll',appraisal.descripcionSector+'\n'+appraisal.descripcionPlanoRegulador+'\n'+appraisal.descripcionExpropiacion)

    ws = wb.worksheets[0]

    for re in appraisal.valuationRealEstate.all():
        excel_find_replace(wb,"propertyAddress",re.addressVerboseNoRegion,once=True)
        excel_find_replace(wb,"terrainSquareMeters",re.house.terrainSquareMeters,once=True)
        excel_find_replace(wb,"builtSquareMeters",re.house.builtSquareMeters,once=True)
        excel_find_replace(wb,"marketPrice",re.marketPrice,once=True)

    excel_find_replace(wb,'propertyAddress','')
    excel_find_replace(wb,'terrainSquareMeters','')
    excel_find_replace(wb,'builtSquareMeters','')
    excel_find_replace(wb,'marketPrice','')

    for i, te in enumerate(realEstate.terrains.all()):
        excel_find_replace(wb,"terrain.name","({}) {}".format(i+1,te.name),once=True)
        excel_find_replace(wb,"terrain.area",te.area,once=True)
        excel_find_replace(wb,"terrain.UFPerArea",te.UFPerArea,once=True)

    excel_find_replace(wb,'terrain.name','')
    excel_find_replace(wb,'terrain.area','')
    excel_find_replace(wb,'terrain.UFPerArea','')

    for i, co in enumerate(realEstate.constructions.all()):
        excel_find_replace(wb,"construction.name","({}) {}".format(i+1,co.name),once=True)
        for a in Construction.MATERIAL_CHOICES:
            if a[0] == co.material:
                excel_find_replace(wb,"construction.material",a[1],once=True)
        excel_find_replace(wb,"construction.year",co.year.year,once=True)
        for a in Construction.BOOLEAN_NULL_CHOICES:
            if a[0] == co.prenda:
                excel_find_replace(wb,"construction.prenda",a[1],once=True)
        for a in Construction.RECEPCION_CHOICES:
            if a[0] == co.recepcion:
                excel_find_replace(wb,"construction.recepcion",a[1],once=True)
        excel_find_replace(wb,"construction.area",co.area,once=True)
        excel_find_replace(wb,"construction.UFPerArea",co.UFPerArea,once=True)

    excel_find_replace(wb,'construction.name','')
    excel_find_replace(wb,'construction.material','')
    excel_find_replace(wb,'construction.year','')
    excel_find_replace(wb,'construction.prenda','')
    excel_find_replace(wb,'construction.recepcion','')
    excel_find_replace(wb,'construction.area','')
    excel_find_replace(wb,'construction.UFPerArea','')

    for i, ass in enumerate(realEstate.assets.all()):
        excel_find_replace(wb,"asset.name","({}) {}".format(i+1,ass.name),once=True)
        excel_find_replace(wb,"asset.value",ass.value,once=True)

    excel_find_replace(wb,'asset.name','')
    excel_find_replace(wb,'asset.value','')

    # Map image
   # r = requests.get("https://maps.googleapis.com/maps/api/staticmap?center="+\
   # 	realEstate.addressStreet+" "+\
   # 	realEstate.addressNumber+"&zoom=16&size=900x530&maptype=roadmap&key=AIzaSyDgwKrK7tfcd9kCtS9RKSBsM5wYkTuuc7E&markers=color:blue%7Clabel:A%7C"+\
    #	str(realEstate.lat)+","+str(realEstate.lng))
    #lf = tempfile.NamedTemporaryFile()
    #lf.write(r.content)
    #img = drawing.image.Image(lf)
    #img.anchor = "B31"
    #wb.worksheets[0].add_image(img)

    photo_coords = ["C210","P210","AE210","AT210"]
    for i, photo in enumerate(appraisal.photos.all()):
        img = drawing.image.Image('uploads/'+photo.photo.name)
        ws.add_image(img)

    response = HttpResponse(
        content=save_virtual_workbook(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="somefilename.xlsx"'

    return response