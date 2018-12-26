from openpyxl import load_workbook
import re
import sys
import os
import django
#sys.path.append('/Users/Pablo Ferreiro/ProyectoInmobiliario/web/') #para pc
sys.path.append('/Users/pabloferreiro/ProyectoInmobiliario/web') #para Mac
os.environ['DJANGO_SETTINGS_MODULE'] = 'map.settings'
django.setup()

from django.core.exceptions import ObjectDoesNotExist
from django.utils.dateparse import parse_date

from realestate.models import RealEstate, Asset
from terrain.models import Terrain
from house.models import House
from building.models import Building
from apartment.models import Apartment
from appraisal.models import Appraisal, Comment, Photo, Document
from commune.models import Commune

from create.create import createOrGetRealEstate

def get_clean_address(rawaddress):
    data = {}
    address = rawaddress.strip()
    n = re.findall('\d+', address)
    regex = r'\b(#\w*[^#\W]|[^#\W]\w*#)\b'.replace('#', "D")
    d = re.split(regex, rawaddress, re.I)
    if len(d) >= 1:
        data['addressNumber2'] = d[-1].strip(". ")
    if len(n) == 3:
        data['addressStreet'] = address.split(n[0])[0].strip().strip("N°n° ")
        data['addressNumber'] = n[0]
        data['addressNumber2'] = d[0]
    elif len(n)== 2:
        data['addressStreet'] = address.split(n[0])[0].strip().strip("N°n° ")
        data['addressNumber'] = n[0]
        data['addressNumber2'] = n[1]

    else:
        print(len(d))
        data['addressStreet'] = address.split(n[0])[0].strip().strip("N°n° ")
        data['addressNumber'] = n[0]
        data['addressNumber2'] = None
        if len(d) > 1:
            data['addressNumber2'] = d[-1].strip(". ")
    if data['addressNumber2']==None:
        data['addressNumber2'] = ' '
    return data

def get_commune_name(commune):
    #gets commune name with first capital letter "la florida" ->"La Florida"
    name = ""
    for word in commune.split(" "):
        name += word.capitalize() + " "
    return name.strip(" ")

def green_stamp(text):
    colors = {
        'Verde':'V',
        'Amarillo':'A', 'Amarillo Vencido':'AV',
        'Rojo':'R',
        'No Aplica':'NA',
        'Verde vencido':'VV', 'Verde Vencido':'VV',
        'Sin antecedentes':'SA',
        'Sin Antecedentes':'SA',
        'No procede': 'SA', 'No encontrado':'SA'}
    return colors[text]

def excel_find_import(workbook1, workbook2, term):
    #finds data of appraisal of Santander xlxs file using santander template file
    for sheet in workbook1.sheetnames:
        for row in range(1, 400):
            for col in range(1, 100):
                cv = workbook1.get_sheet_by_name(sheet).cell(row=row, column=col).value
                if cv == term:
                    value = workbook2.get_sheet_by_name(sheet).cell(row=row, column=col).value
                    if value == None:
                        value = workbook2.get_sheet_by_name(sheet).cell(row=row+1, column=col).value
                        return value
                    else:
                        return value

def excel_find_general(file, term):
    # finds data by term in appraisal file
    print(term)
    wb = load_workbook(filename=file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    if term == "PROPIEDAD ANALIZADA":
        for row in range(120, 200):
            for col in range(1, 25):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print('Found it')
                    terrainSquareMeters = ws.cell(row=row, column=col + 24).value
                    builtSquareMeters = ws.cell(row=row, column=col + 30).value
                    return terrainSquareMeters, builtSquareMeters
    elif term == "VALOR COMERCIAL":
        for row in range(74, 120):
            for col in range(20, 60):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    col = 54
                    valorUF = round(ws.cell(row=row, column=col).value,2)
                    return valorUF
    elif term == "Antigüedad" or term == "Vida Util" or term == "Sello de Gases" or term == "Tipo Propiedad":
        for row in range(35, 70):
            for col in range(1, 30):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    col = 6
                    feature = ws.cell(row=row, column=col).value
                    return feature
    elif term == "Total Avalúo Fiscal" or term == "N° Rol Principal" or term== "N° Rol (es) Sec.":
        for row in range(35, 70):
            for col in range(10, 30):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    col = 17
                    avaluo = ws.cell(row=row, column=col).value
                    return avaluo
    elif term == "Leyes que se Acoge":
        print(term)
        for row in range(35, 70):
            for col in range(5, 20):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    col = 10
                    ley1 = ws.cell(row=row, column=col).value
                    ley2 = ws.cell(row=row + 1, column=col).value
                    return ley1, ley2
    elif term == "Permiso Edificación":
        print(term)
        for row in range(35, 70):
            for col in range(5, 20):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    col = 10
                    permiso = ws.cell(row=row, column=col).value
                    fecha  = ws.cell(row=row, column=col + 1).value
                    return permiso, fecha
    elif term == "II. DESCRIPCIÓN GENERAL DEL BIEN TASADO":
        print(term)
        for row in range(14, 20):
            for col in range(1, 10):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    descripcion = ws.cell(row=row + 1, column=col).value
                    return descripcion
    elif term == "Sub Total Terreno" or term == "Sub Total Construcciones":
        print(term)
        for row in range(40, 70):
            for col in range(8, 15):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    metros = ws.cell(row=row, column=col - 4).value
                    return metros
    elif term == "Valor Comercial":
        print(term)
        for row in range(85, 100):
            for col in range(11, 20):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    valorUf = ws.cell(row=row, column=col + 3).value
                    liquidacion = ws.cell(row=row + 1, column=col + 2).value
                    return valorUf, liquidacion

def findFromDescription(text):
    baños = 0
    dormitorio = 0
    for i in range(len(text.split(' '))):
        word = text.split(' ')[i]
        if word == 'baño' or word == 'Baño' or word == 'baños' or word == 'Baños':
            try:
                baños += int(text.split(' ')[i+1])
            except ValueError:
                continue
            try:
                baños += int(text.split(' ')[i-1])
            except ValueError:
                continue

            print(word)
            print(text.split(' ')[i+1])
            print(text.split(' ')[i-1])
    print(baños)


def importAppraisalSantander(file):

    def convert(tude):
        #convert degrees, minutes, secods coordiantes into decimals

        try:
            multiplier = 1 if tude[-1] in ['N', 'E'] else -1
            d = float(tude[:-1].split('°')[0].replace(',','.'))
            m = float(tude[:-1].split('°')[1].split("'")[0].replace(',','.'))/60
            s =tude[:-1].split('°')[1].split("'")[1].split("''")[0].replace(',','.')
            s = re.sub('[^\d\.]', '', s)
            s = float(s)/3600
            print((d + m + s)*multiplier)
            return (d + m + s)*multiplier
        except ValueError:
            return tude


    def ocupantes_choices(ocupante):
        #converts from file format to database format
        ocupante = ocupante.capitalize()
        if ocupante=='Propietario':
            ocupanteFinal='P'
        elif ocupante=='Arrendatario':
            ocupanteFinal='A'
        elif ocupante=='Sin ocupante':
            ocupanteFinal='SO'
        else:
            ocupanteFinal='O'
        return ocupanteFinal

    def destinoSII_modified(destino):
        # converts from file format to database format
        listaSII= {'H-Habitacional':'H','HABITACIONAL':'H',
                'O-Oficina':'O', 'OFICINA':'O',
                'C-Comercio':'C', 'COMERCIO':'C', 'COMERCIAL':'C',
                'I-Industria':'I', 'INDUSTRIA':'I',
                'L-Bodega':'L', 'BODEGA':'L',
                'Z-Estacionamiento':'Z', 'ESTACIONAMIENTO':'Z',
                'D-Deportes y Recreación':'D','DEPORTES Y RECREACIÓN':'D',
                'E-Educación y Cultura':'E','EDUCACIÓN Y CULTURA':'E',
                'G-Hotel, Motel':'G','HOTEL, MOTEL':'G',
                'P-Administración pública':'P','ADMINISTRACIÓN PÚBLICA':'P',
                'Q-Culto':'Q', 'CULTO':'Q',
                'S-Salud':'S', 'SALUD':'S',
                'SITIO ERIAZO': 'SO',
                'PROFESIONAL (oficina)':'O'   }
        return(listaSII[destino])

    def date_to_datetimefield(rawdate):
        try:
            finalDate = rawdate.split('//')[-1].strip(' ').split('-')[-1] + '-' + \
                                 rawdate.split('//')[-1].strip(' ').split('-')[-2] + \
                                 '-' + rawdate.split('//')[-1].strip(' ').split('-')[-3]
        except IndexError:
            finalDate = "1" + "-" + "1" + "-" +rawdate.split(' ')[-1]
        return parse_date(finalDate)

    def boolean_null_choices(choice):
        options={
            "S/A":1, "S/Ant.":1,
            "Si":2, "SI":2,
            "No":3, "NO":3,
            "No Aplica":1, " ":1, "":1}
        return options[choice]

    def law_to_database(text):
        law ={
            'O.G.U. y C.':0,
            'P.R.C.':1,
            'Ley Pereira':2,
            'Ley 19583': 3,
            'Ley 19667':4,
            'Ley 19727':5,
            'Ley 20251':6,
            'Ley 6071':7, 'Ley N° 6071':7,
            'Ninguna':8,
            'Antigüedad':9
        }
        return law[text]

    module_dir = os.path.dirname(__file__)  # get current directory
    file_path = os.path.join(module_dir, 'static/appraisal/santander-template.xlsx')
    wb = load_workbook(filename=file_path)
    wb2 = load_workbook(filename=file, read_only=True, data_only=True)

    solicitanteCodigo = excel_find_import(wb, wb2, "solicitanteCodigo")
    id = excel_find_import(wb, wb2, "id")
    timeModified = excel_find_import(wb, wb2, "timeModified")
    solicitanteSucursal = excel_find_import(wb, wb2, "solicitanteSucursal")
    solicitanteEjecutivo = excel_find_import(wb, wb2, "solicitanteEjecutivo")
    rol = excel_find_import(wb, wb2, "rol")
    cliente = excel_find_import(wb, wb2, "cliente")
    clienteRut = excel_find_import(wb, wb2, "clienteRut")
    propietario = excel_find_import(wb, wb2, "propietario")
    propietarioRut = excel_find_import(wb, wb2, "propietarioRut")
    address = get_clean_address(excel_find_import(wb, wb2, "address"))
    addressCommune = get_commune_name(excel_find_import(wb, wb2, "addressCommune"))
    addressRegion = excel_find_import(wb, wb2, "addressRegion")
    #tasadorUser = excel_find_import(wb, wb2, "tasadorUser")
    #tasadorUserrut = excel_find_import(wb, wb2, "tasadorUser.rut")
    lat = convert(excel_find_import(wb, wb2, "lat"))
    lng = convert(excel_find_import(wb, wb2, "lng"))
    mercadoObjetivo = boolean_null_choices(excel_find_import(wb, wb2, "mercadoObjetivo"))
    antiguedad = excel_find_import(wb, wb2, "antiguedad")
    vidaUtil = excel_find_import(wb, wb2, "vidaUtil")
    if vidaUtil != int or float:
        vidaUtil = 0
    avaluoFiscal = excel_find_import(wb, wb2, "avaluoFiscal")
    acogidaLey = law_to_database(excel_find_import(wb, wb2, "acogidaLey"))
    dfl2 = boolean_null_choices(excel_find_import(wb, wb2, "dfl2"))
    selloVerde = green_stamp(excel_find_import(wb, wb2, "selloVerde"))
    copropiedadInmobiliaria = boolean_null_choices(excel_find_import(wb, wb2, "copropiedadInmobiliaria"))
    ocupante = ocupantes_choices(excel_find_import(wb, wb2, "ocupante"))
    tipoBien = excel_find_import(wb, wb2, "tipoBien")
    destinoSII = destinoSII_modified(excel_find_import(wb, wb2, "destinoSII"))
    usoActual = destinoSII_modified(excel_find_import(wb, wb2, "usoActual"))
    usoFuturo = destinoSII_modified(excel_find_import(wb, wb2, "usoFuturo"))
    permisoEdificacion = excel_find_import(wb, wb2, "permisoEdificacion")
    recepcionFinal = excel_find_import(wb, wb2, "recepcionFinal")
    expropiacion = boolean_null_choices(excel_find_import(wb, wb2, "expropiacion"))
    viviendaSocial = boolean_null_choices(excel_find_import(wb, wb2, "viviendaSocial"))
    adobe = boolean_null_choices(excel_find_import(wb, wb2, "adobe"))
    desmontable = boolean_null_choices(excel_find_import(wb, wb2, "desmontable"))
    generalDescription = excel_find_import(wb, wb2, "generalDescription")
    descripcionSectorAll = excel_find_import(wb, wb2, "descripcionSectorAll")
    programa = excel_find_import(wb, wb2, "programa")
    estructuraTerminaciones = excel_find_import(wb, wb2, "estructuraTerminaciones")
    mm2 = excel_find_general(file, "PROPIEDAD ANALIZADA")
    terrainSquareMeters = mm2[0]
    usefulSquareMeters = mm2[0]
    builtSquareMeters = mm2[1]
    terraceSquareMeters = mm2[1]
    valorUF = excel_find_general(file, "VALOR COMERCIAL")

    #hardcoded for now
    ws2 = wb2.worksheets[0]
    propertyType = ws2['U5'].value

    # Crear Realestate

    propiedad = createOrGetRealEstate(
        addressNumber=address['addressNumber'],
        addressStreet=address['addressStreet'],
        addressCommune=Commune.objects.get(name=addressCommune),
        addressRegion=Commune.objects.get(name=addressCommune).region)
    propiedad.lat = lat
    propiedad.lng = lng
    print(propiedad)

    # Crear building
    edificio = propiedad.createOrGetEdificio(addressNumber2=address['addressNumber2'])
    edificio.propertyType = propertyType
    edificio.name = str(file)
    # edificio.marketPrice = valorUF
    edificio.vidaUtilRemanente = vidaUtil
    edificio.dfl2 = dfl2
    edificio.avaluoFiscal = avaluoFiscal
    edificio.copropiedadInmobiliaria = copropiedadInmobiliaria
    edificio.selloVerde = selloVerde
    edificio.permisoEdificacionNo = permisoEdificacion
    edificio.permisoEdificacionFecha = recepcionFinal
    edificio.tipoPropiedad = tipoBien
    edificio.rol = rol
    edificio.year = recepcionFinal
    edificio.mercadoObjetivo = mercadoObjetivo
    edificio.antiguedad = antiguedad
    edificio.acogidaLey = acogidaLey
    edificio.ocupante = ocupante
    edificio.adobe = adobe
    edificio.expropiacion = expropiacion
    edificio.viviandaSocial = viviendaSocial
    edificio.desmontable = desmontable
    edificio.usoActual = usoActual
    edificio.usoFuturo = usoFuturo
    edificio.destinoSII = destinoSII

    edificio.save()

    if propertyType == Building.TYPE_CASA:
        casa = propiedad.createOrGetCasa(addressNumber2=address['addressNumber2'])
        # casa.bedrooms = bedrooms
        # casa.bathrooms = bathrooms
        casa.builtSquareMeters = builtSquareMeters
        casa.terrainSquareMeters = terrainSquareMeters
        casa.generalDescription = generalDescription
        casa.marketPrice = valorUF

        casa.save()
    elif propertyType == Building.TYPE_DEPARTAMENTO:
        departamento = propiedad.createOrGetDepartamento(addressNumber2=address['addressNumber2'])
        # departamento.floor = floor
        # departamento.orientation = orientation
        # departamento.bedrooms = bedrooms
        # departamento.bathrooms = bathdrooms
        departamento.usefulSquaremeters = usefulSquareMeters
        departamento.terraceSquaremeters = terraceSquareMeters
        departamento.generalDescription = generalDescription
        departamento.marketPrice = valorUF
        departamento.programa = programa
        departamento.save()

    elif propertyType == Building.TYPE_TERRENO:
        terreno = Terrain(name=str(file), area=terrainSquareMeters, rol=rol)
        terreno.marketPrice = valorUF
        terreno.save()
        propiedad.terrains = terreno
        propiedad.save()

    print(propertyType)

    # crear tasación

    try:
        appraisal = Appraisal.objects.get(real_estates=propiedad,
                                          tipoTasacion=1,
                                          state=0,
                                          source=1,
                                          solicitanteCodigo=solicitanteCodigo,
                                          timeFinished=timeModified
                                          )
        appraisal.solicitante = 2
        appraisal.visita = 1
        appraisal.solicitanteOtro = id
        appraisal.solicitanteEjecutivo = solicitanteEjecutivo
        appraisal.solicitanteSucursal = solicitanteSucursal
        appraisal.cliente = cliente
        appraisal.clienteRut = clienteRut
        appraisal.propietario = propietario
        appraisal.propietarioRut = propietarioRut
        appraisal.descripcionSector = descripcionSectorAll
        #appraisal.liquidez = valorLiquidez
        #appraisal.tasadorUser = tasadorUser
        appraisal.timeFinished = timeModified
        appraisal.save()
        print('Existe')

    except ObjectDoesNotExist:
        appraisal = Appraisal(state=0,
                              source=1,
                              tipoTasacion=1,
                              solicitante=2,
                              visita=1,
                              solicitanteCodigo=solicitanteCodigo,
                              solicitanteOtro=id,
                              timeFinished=timeModified,
                              solicitanteEjecutivo=solicitanteEjecutivo,
                              solicitanteSucursal=solicitanteSucursal,
                              cliente=cliente,
                              clienteRut=clienteRut,
                              propietario=propietario,
                              propietarioRut=propietarioRut,
                              descripcionSector=descripcionSectorAll,
                              #liquidez=valorLiquidez
                              #tasadorUser=tasadorUser,
                              )

        appraisal.save()
        appraisal.real_estates.add(propiedad)
        appraisal.save()
        print('No Existe')


def importAppraisalITAU(file):

    def tipoPropiedad(text):
        tipos ={
            'Casa': Building.TYPE_CASA,
            'Casa en Parcela de Agrado': Building.TYPE_PARCELA,
            'Departamento': Building.TYPE_DEPARTAMENTO,
            'Terreno Unifamiliar': Building.TYPE_TERRENO,
        }
        return tipos[text]

    def law_to_database(text, text2, law):
        if law == "DFL2" and text == "DFL. 2 /59 (Viv. Económica)" or text2 == "DFL. 2 /59 (Viv. Económica)":
            return 2
        elif law == 'copropiedad' and text == "L. 19537 /97 (Cop. Inmobiliaria)" or text2 == "L. 19537 /97 (Cop. Inmobiliaria)":
            return 2
        else:
            return 0

    def estadoPropiedad(text):
        estado = {
            'Nueva':0,
            'Usada':1,
            'No Aplica':3,
            'Sitio Eriazo':3
        }
        return estado[text]


    module_dir = os.path.dirname(__file__)  # get current directory
    file_path = os.path.join(module_dir, 'static/appraisal/itau-template.xlsx')
    wb = load_workbook(filename=file_path)
    wb2 = load_workbook(filename=file, read_only=True, data_only=True)

    solicitanteCodigo = excel_find_import(wb, wb2, "solicitanteCodigo")
    id = excel_find_import(wb, wb2, "id")
    solicitanteEjecutivo = excel_find_import(wb, wb2, "solicitanteEjecutivo")
    cliente = excel_find_import(wb, wb2, "cliente")
    clienteRut = str(excel_find_import(wb, wb2, "clienteRut")) + '-' +  str(excel_find_import(wb, wb2, "num1"))
    propietario = excel_find_import(wb, wb2, "propietario")
    propietarioRut = str(excel_find_import(wb, wb2, "propietarioRut")) + '-' +  str(excel_find_import(wb, wb2, "num1"))
    addressStreet = excel_find_import(wb, wb2, "addressStreet")
    addressNumber = excel_find_import(wb, wb2, "addressNumber")
    addressNumber2 = excel_find_import(wb, wb2, "addressNumber2")
    addressCommune = get_commune_name(excel_find_import(wb, wb2, "addressCommune"))
    addressRegion = excel_find_import(wb, wb2, "addressRegion")
    rol1 = excel_find_general(file, "N° Rol Principal")
    fechaVisita = excel_find_import(wb, wb2, "timeModified")
    print(fechaVisita)
    rol2 = excel_find_general(file, "N° Rol (es) Sec.")
    #tasadorUser = excel_find_import(wb, wb2, "tasadorUser")
    # lat = convert(excel_find_import(wb, wb2, "lat")) #Itau no viene con lat-long, usar función?
    # lng = convert(excel_find_import(wb, wb2, "lng"))
    antiguedad = int(excel_find_general(file, "Antigüedad"))
    vidaUtil = int(excel_find_general(file, "Vida Util"))
    vidaUtilRemanente = vidaUtil-antiguedad
    avaluoFiscal = excel_find_general(file, "Total Avalúo Fiscal")
    leyes = excel_find_general(file, "Leyes que se Acoge")
    acogidaLey = leyes[0]
    acogidaLey2 = leyes[1]
    selloVerde = green_stamp(excel_find_general(file, "Sello de Gases"))
    tipoBien = estadoPropiedad(excel_find_general(file, "Tipo Propiedad"))
    permiso = excel_find_general(file, "Permiso Edificación")
    permisoEdificacion = permiso[0]
    recepcionFinal = permiso[1]
    generalDescription = excel_find_general(file, "II. DESCRIPCIÓN GENERAL DEL BIEN TASADO")
    terrainSquareMeters = excel_find_general(file, "Sub Total Terreno")
    builtSquareMeters = excel_find_general(file, "Sub Total Construcciones")
    propertyType = tipoPropiedad(excel_find_import(wb, wb2, "propertyType"))
    valores = excel_find_general(file, "Valor Comercial")
    valorUF = valores[0]
    valorLiquidez = valores[1]
    dfl2 = law_to_database(acogidaLey, acogidaLey2, "DFL2")
    copropiedadInmobiliaria = law_to_database(acogidaLey, acogidaLey2, "copropiedad")


    #Crear Realestate


    propiedad = createOrGetRealEstate(
        addressNumber=addressNumber,
        addressStreet=addressStreet,
        addressCommune=Commune.objects.get(name=addressCommune),
        addressRegion=Commune.objects.get(name=addressCommune).region)
    print(propiedad)

    #Crear building
    edificio = propiedad.createOrGetEdificio(addressNumber2=addressNumber2)
    edificio.propertyType = propertyType
    edificio.name = str(file)
    #edificio.marketPrice = valorUF
    edificio.vidaUtilRemanente = vidaUtilRemanente
    edificio.dfl2 = dfl2
    edificio.avaluoFiscal = avaluoFiscal
    edificio.copropiedadInmobiliaria = copropiedadInmobiliaria
    edificio.selloVerde = selloVerde
    edificio.permisoEdificacionNo = permisoEdificacion
    edificio.permisoEdificacionFecha = recepcionFinal
    edificio.tipoPropiedad = tipoBien
    edificio.rol = rol1
    edificio.year = recepcionFinal

    edificio.save()


    if propertyType == Building.TYPE_CASA:
        casa = propiedad.createOrGetCasa(addressNumber2=addressNumber2)
        #casa.bedrooms = bedrooms
        #casa.bathrooms = bathrooms
        casa.builtSquareMeters = builtSquareMeters
        casa.terrainSquareMeters = terrainSquareMeters
        casa.generalDescription = generalDescription
        casa.marketPrice = valorUF

        casa.save()
    elif propertyType == Building.TYPE_DEPARTAMENTO:
        departamento = propiedad.createOrGetDepartamento(addressNumber2=addressNumber2)
        #departamento.floor = floor
        #departamento.orientation = orientation
        #departamento.bedrooms = bedrooms
        #departamento.bathrooms = bathdrooms
        departamento.usefulSquaremeters = builtSquareMeters
        departamento.generalDescription = generalDescription
        departamento.marketPrice = valorUF
        departamento.save()

    elif propertyType == Building.TYPE_TERRENO:
        terreno = Terrain(name=str(file), area=terrainSquareMeters, rol=rol1)
        terreno.marketPrice = valorUF
        terreno.save()
        propiedad.terrains=terreno
        propiedad.save()

    print(propertyType)

    # crear tasación

    try:
        appraisal = Appraisal.objects.get(real_estates=propiedad,
                                          tipoTasacion=1,
                                          state=0,
                                          source=1,
                                          solicitanteCodigo=solicitanteCodigo,
                                          timeFinished=fechaVisita
                                          )
        appraisal.solicitante = 3
        appraisal.solicitanteOtro = id
        appraisal.solicitanteEjecutivo = solicitanteEjecutivo
        appraisal.cliente = cliente
        appraisal.clienteRut = clienteRut
        appraisal.propietario = propietario
        appraisal.propietarioRut = propietarioRut
        appraisal.visita = 1
        #appraisal.liquidez = valorLiquidez
        # appraisal.tasadorUser=tasadorUser
        appraisal.save()
        print('Existe')

    except ObjectDoesNotExist:
        appraisal = Appraisal(state=0,
                              source=1,
                              tipoTasacion=1,
                              solicitante=3,
                              visita=1,
                              solicitanteCodigo=solicitanteCodigo,
                              solicitanteOtro=id,
                              timeFinished=fechaVisita,
                              solicitanteEjecutivo=solicitanteEjecutivo,
                              cliente=cliente,
                              clienteRut=clienteRut,
                              propietario=propietario,
                              propietarioRut=propietarioRut,
                              #liquidez=valorLiquidez
                              #tasadorUser=tasadorUser,
                              )

        appraisal.save()
        appraisal.real_estates.add(propiedad)
        appraisal.save()
        print('No Existe')





file1 = 'TMI 1803234 Antonio Patricio Moder Donoso (13566161-9) Independencia 1142 Casa 4 Condominio Parque Don Antonio Puente Alto mod 23-10-18.xlsx'
file2 = 'TMI-1803741 Michael Alarcon Fernandez (13685545-K) Lago Hurón 1444 Villa Canadá Maipú inc 24-10-18.xlsx'
file3 = 'TMI-1805225 Gonzalo Garrido (13678613-K) Doctor Johow 550 Departamento 44-D Bloque D Conjunto Dr Johow Ñuñoa.xlsx'
file_mac = '/Volumes/GoogleDrive/Mi unidad/ProyectoInmobiliario/Datos/tasaciones/'
file_pc = 'G:/Mi unidad/ProyectoInmobiliario/Datos/tasaciones/'
file = file_mac + file1
#importAppraisalSantander(file)
#importAppraisalITAU(file)

text = "Se analiza departamento ubicado en 1° piso, con vista al poniente, con vista a calle aledaña. Cuenta con un programa arquitectónico consiste en estar-comedor, cocina, logia, baño y 2 dormitorios. Mantiene nivel estándar de terminaciones y buen estado de conservación. Inmueble no incorpora obras complementarias. Copropiedad cuenta con bloques de edificio destinados a departamentos habitación. La unidad tasada pertenece a edificio de 4 pisos, sin subterráneo. Se abastece sólo con caja de escala. Copropiedad no cuenta con equipamiento comunitario, según datos aportados en visita. El sector corresponde a área cercana a A. Libertador Bdo. O'Higgins, Las Rejas y General Velásquez, principales ejes estructurantes dentro de la comuna y su entorno, orientada a segmentos socioeconómicos medios, conformada además por conjunto de edificios de departamentos de igual altura en misma copropiedad y sector. Cercano a Estaciones de Metro: Ecuador y Las Rejas. Posee buena accesibilidad  y amplio equipamiento de apoyo dado su emplazamiento.  Municipalidad de Estación Central señala verbalmente que se el inmueble posee P.E. N°06/88 del año 1988 y R.F. N°59 de fecha 26 de octubre de 1988 por una superficie de 46,25 m2. Se acoge a DFL N°2 de 1959, D.L. N°2552 DE 1979 y Ley N°6.071. Plano de Loteo aprobado mediante Res. N°16 de fecha 21 de septiembre de 1988."

findFromDescription(text)