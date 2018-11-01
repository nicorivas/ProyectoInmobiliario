from django.core import serializers
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.core.exceptions import MultipleObjectsReturned

from realestate.models import RealEstate
from house.models import House
from building.models import Building
from apartment.models import Apartment
from appraisal.models import Appraisal
from appraisal.models import Comment
from appraisal.models import Photo

import reversion
from copy import deepcopy
from reversion.models import Version

import os
import csv

from .forms import FormBuilding
from .forms import FormApartment
from .forms import FormAppraisal
from .forms import FormComment
from .forms import FormHouse
from .forms import FormPhotos

from django.db.models import Avg, StdDev

import json

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook

import datetime

def get_realestate(request,id):
    '''
        Given building id, returns the building object.
        It checks for some errors and sends to the error page.
    '''
    realestate = RealEstate.objects.filter(id=id)
    # This must be only one building
    if len(realestate) == 0:
        context = {'error_message': 'Real estate should exist by now'}
        return render(request, 'appraisal/error.html',context)
    elif len(realestate) > 1:
        context = {'error_message': 'Se encontró más de una propiedad (error base!)'}
        return render(request, 'appraisal/error.html',context)
    realestate = realestate[0]
    return realestate

def get_building(request,id):
    '''
        Given building id, returns the building object.
        It checks for some errors and sends to the error page.
    '''
    building = Building.objects.filter(id=id)
    # This must be only one building
    if len(building) == 0:
        context = {'error_message': 'Building should exist by now'}
        return render(request, 'appraisal/error.html',context)
    elif len(building) > 1:
        context = {'error_message': 'Se encontró más de una propiedad (error base!)'}
        return render(request, 'appraisal/error.html',context)
    building = building[0]
    return building

def get_house(request,id):
    '''
        Given building id, returns the building object.
        It checks for some errors and sends to the error page.
        TODO: Change this to proper try statements
    '''
    house = House.objects.filter(id=id)
    # This must be only one building
    if len(house) == 0:
        context = {'error_message': 'House should exist by now'}
        return render(request, 'appraisal/error.html',context)
    elif len(house) > 1:
        context = {'error_message': 'Se encontró más de una propiedad (error base!)'}
        return render(request, 'appraisal/error.html',context)
    house = house[0]
    return house

def get_apartment(request,id):
    '''
        Given an appartment id, return the apartment object.
        It checks for errors and sends to the correct error page.
    '''
    apartment = Apartment.objects.filter(id=id)
    # This must be only one apartment
    if len(apartment) == 0:
        context = {'error_message': 'Apartment should exist by now'}
        return render(request, 'appraisal/error.html',context)
    elif len(apartment) > 1:
        context = {'error_message': 'Se encontró más de una propiedad (error base!)'}
        return render(request, 'appraisal/error.html',context)
    apartment = apartment[0]
    return apartment

def get_appraisal(request,id):
    '''
        Given an appraisal id, returns the apraisal object.
        It checks for errors and sends to the correct error page.
    '''
    try:
        appraisal = Appraisal.objects.get(pk=id)
        appraisal = appraisal
        return appraisal
    except Appraisal.DoesNotExist:
        context = {'error_message': 'Appraisal not found?'}
        return render(request, 'appraisal/error.html', context)
    except MultipleObjectsReturned:
        context = {'error_message': 'More than one appraisal of the same property'}
        return render(request, 'appraisal/error.html', context)

def get_similar_realestate(realestate):

    if realestate.propertyType == RealEstate.TYPE_APARTMENT:
        if (realestate.apartment.bedrooms != None and
            realestate.apartment.bathrooms != None and
            realestate.apartment.builtSquareMeters != None and
            realestate.apartment.usefulSquareMeters != None):

            apartments = Apartment.objects.filter(
                bedrooms=realestate.apartment.bedrooms,
                bathrooms=realestate.apartment.bathrooms,
                marketPrice__isnull=False).exclude(marketPrice=0)

            ds = []
            ni = 0
            for i, apt in enumerate(apartments):
                d1 = float(pow(realestate.apartment.builtSquareMeters - apt.builtSquareMeters,2))
                d2 = float(pow(realestate.apartment.usefulSquareMeters - apt.usefulSquareMeters,2))
                ds.append([0,0])
                ds[i][0] = apt.pk
                ds[i][1] = d1+d2

            ds = sorted(ds, key=lambda x: x[1])
            ins = [x[0] for x in ds]

            references = realestate.apartments.filter(pk__in=ins[0:5])
            return references
        else:
            return []
    else:
        return []


def save_appraisal(request,forms,comment):
    print('save_appraisal')
    with reversion.create_revision():
        print(forms['appraisal'].cleaned_data)
        forms['appraisal'].save()
        reversion.set_user(request.user)
        reversion.set_comment(comment)
        print('a')
        return

def save(request,forms,realEstate):
    print('save')
    if realEstate.propertyType == RealEstate.TYPE_APARTMENT:
        if forms['building'].is_valid() and forms['apartment'].is_valid() and forms['appraisal'].is_valid():
            print('holi')
            for name, form in forms.items():
                if name in ['building','apartment']:
                    form.save()
                if name == 'appraisal':
                    save_appraisal(request,forms,'Saved')
            return True
        else:
            print(forms['building'].errors)
            print(forms['apartment'].errors)
            print(forms['appraisal'].errors)
    elif realEstate.propertyType == RealEstate.TYPE_HOUSE:
        if forms['house'].is_valid() and forms['appraisal'].is_valid():
            for name, form in forms.items():
                if name == 'house':
                    form.save()
                if name == 'appraisal':
                    save_appraisal(request,forms,'Saved')
            return True
    else:
        return False

def delete(request,appraisal):
    appraisal.delete()
    context = {}
    return render(request,'appraisal/deleted.html',context)

def finish(request,forms, appraisal):
    for name, form in forms.items():
        if not form.is_valid():
            print(form.errors)
            return False
    appraisal.timeFinished = datetime.datetime.now()
    appraisal.state = Appraisal.STATE_FINISHED
    save_appraisal(request,forms,'Finished')
    print('finish')
    return False

def comment(forms,appraisal):
    '''
    Create comment based on the field commentText of the form.
    '''
    if forms['comment'].is_valid():
        text = forms['comment'].cleaned_data['commentText']
        conflict = forms['comment'].cleaned_data['commentConflict']
        if conflict:
            appraisal.state = appraisal.STATE_PAUSED
            appraisal.timePaused = datetime.datetime.now()
            appraisal.save()
        comment = Comment(
            user=request.user,
            text=text,
            timeCreated=datetime.datetime.now(),
            appraisal=appraisal,
            conflict=conflict)
        comment.save()
    return True

def export(forms):
    module_dir = os.path.dirname(__file__)  # get current directory
    file_path = os.path.join(module_dir,'static/appraisal/test.xlsx')

    def excel_find(workbook,term):
        for sheet in workbook:
            for row in range(1,100):
                for col in range(1,100):
                   cv = sheet.cell(row=row,column=col).value

    def excel_find_replace(workbook,term,rep):
        for sheet in workbook:
            for row in range(1,100):
                for col in range(1,100):
                   cv = sheet.cell(row=row,column=col).value
                   if cv == term:
                       sheet.cell(row=row,column=col).value = rep

    wb = load_workbook(filename=file_path)

    model_instance = form_appraisal.save(commit=False)

    fields = Appraisal._meta.get_fields()
    for field in fields:
        field_name = field.deconstruct()[0]
        excel_find_replace(wb,field_name,getattr(model_instance,field_name))

    response = HttpResponse(
        content=save_virtual_workbook(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="somefilename.xlsx"'

    return response

def assign_visador(request,forms,appraisal):
    if forms['appraisal'].is_valid():
        appraisal.visadorUser = User.objects.get(pk=request.POST.dict()['visador'])
        save_appraisal(request, forms, 'Changed visador')
        return True

def assign_tasador(request,forms,appraisal):
    if forms['appraisal'].is_valid():
        appraisal.tasadorUser = User.objects.get(pk=request.POST.dict()['tasador'])
        save_appraisal(request, forms, 'Changed tasador')
        return True

def upload_photo(request,forms,appraisal):
    if forms['photos'].is_valid() and forms['appraisal'].is_valid():
        for photo_file in request.FILES.getlist('photos'):
            photo = Photo()
            photo.photo = photo_file
            photo.description = forms['photos'].cleaned_data['description']
            photo.save()
            appraisal.photos.add(photo)
        save_appraisal(request, forms, 'Added picture(s)')

def delete_photo(request,forms,appraisal):
    appraisal.photos.remove(request.POST['photo_id'])
    if forms['appraisal'].is_valid():
        save_appraisal(request, forms, 'Removed picture(s)')

def appraisal(request, **kwargs):
    '''
    General view for appraisals. Gets a variable number of parameters depending
    on the type of realestate.
    '''

    # Get current appraisal
    appraisal = get_appraisal(request,kwargs['appraisal_id'])
    if isinstance(appraisal,HttpResponse): return appraisal
    appraisal_old = deepcopy(appraisal) # done to check differences when saving history

    # Get realestate
    realestate = get_realestate(request,appraisal.realEstate.id)
    if realestate == None:
        context = {'error_message': 'Realestae was not found'}
        return render(request, 'appraisal/error.html',context)

    # DIE FORMS

    # if this is a POST request we need to process the form data
    if request.method == 'POST':

        print(request.method)

        # Process forms
        forms = {}
        forms['appraisal'] = FormAppraisal(request.POST,request.FILES,instance=appraisal)
        forms['comment'] = FormComment(request.POST)
        if realestate.propertyType == RealEstate.TYPE_APARTMENT:
            forms['apartment'] = FormApartment(request.POST,instance=realestate.apartment)
            forms['building'] = FormBuilding(request.POST,instance=realestate.apartment.building_in)
        elif realestate.propertyType == RealEstate.TYPE_HOUSE:
            forms['house'] = FormHouse(request.POST,instance=realestate.house)
        forms['photos'] = FormPhotos(request.POST,request.FILES)

        print(request.POST.keys())

        # Switch to action
        if 'btn_save' in request.POST.keys():
            ret = save(request,forms,realestate)
        elif 'btn_delete' in request.POST.keys():
            ret = delete(request,appraisal)
        elif 'btn_finish' in request.POST.keys():
            ret = finish(request,forms,appraisal)
        elif 'btn_comment' in request.POST.keys():
            ret = comment(forms,appraisal)
        elif 'btn_assign_tasador' in request.POST.keys():
            ret = assign_tasador(request,forms,appraisal)
        elif 'btn_assign_visador' in request.POST.keys():
            ret = assign_visador(request,forms,appraisal)
        elif 'btn_upload_photo' in request.POST.keys():
            ret = upload_photo(request,forms,appraisal)
        elif 'btn_delete_photo' in request.POST.keys():
            print('btn_delete_photo')
            ret = delete_photo(request,forms,appraisal)
        else:
            ret = False

        if isinstance(ret, HttpResponse): return ret

    # REFERENCE PROPERTIES
    references = []#get_similar_realestate(realestate)

    if len(references) > 0:
        averages = references.aggregate(
            Avg('marketPrice'),
            Avg('builtSquareMeters'),
            Avg('usefulSquareMeters'))
        stds = references.aggregate(
            StdDev('marketPrice'),
            StdDev('builtSquareMeters'),
            StdDev('usefulSquareMeters'))
    else:
        averages = []
        stds = []

    # Visadores and tasadores for the Bootstrap modals where you can select
    # them.
    tasadores = User.objects.filter(groups__name__in=['tasador'])
    visadores = User.objects.filter(groups__name__in=['visador'])

    #Data from de creation form
    tipoTasacion = appraisal.tipoTasacion
    objetivo = appraisal.objetivo
    solicitante = appraisal.solicitante
    print(appraisal.solicitante)

    # History of changes, for the logbook
    versions = list(Version.objects.get_for_object(appraisal))
    appraisal_history = []
    c = 0

    for i in range(len(versions)):
        if i == 0: continue
        version_new = versions[i]
        version_old = versions[i-1]

        appraisal_history.append({})
        appraisal_history[c]['user'] = version_new.revision.user
        appraisal_history[c]['time'] = version_new.revision.date_created
        appraisal_history[c]['diffs'] = []
        for key, value in version_new.field_dict.items():
            if value != version_old.field_dict[key]:
                appraisal_history[c]['diffs'].append({
                    'verbose_name':Appraisal._meta.get_field(key).verbose_name,
                    'name':key,
                    'value':version_old.field_dict[key],
                    'value_old':value})
        if len(appraisal_history[c]['diffs']) == 0:
            appraisal_history.pop(c)
        else:
            c += 1

    # Comments, for the logbook
    comments = Comment.objects.filter(appraisal=appraisal).order_by('-timeCreated')

    # Forms
    forms = {
        'appraisal': FormAppraisal(instance=appraisal,label_suffix=''),
        'comment':FormComment(label_suffix=''),
        'photos':FormPhotos(label_suffix='')}
    if realestate.propertyType == RealEstate.TYPE_APARTMENT:
        forms['realestate'] = FormApartment(instance=realestate.apartment,label_suffix='')
        forms['building'] = FormBuilding(instance=realestate.apartment.building_in,label_suffix='')
    elif realestate.propertyType == RealEstate.TYPE_HOUSE:
        forms['realestate'] = FormHouse(instance=realestate.house,label_suffix='')

    # Disable fields if appraisal is finished
    if appraisal.state == appraisal.STATE_FINISHED or appraisal.state == appraisal.STATE_PAUSED:
        for key, form in forms.items():
            for field in form.fields:
                form.fields[field].widget.attrs['readonly'] = True
                form.fields[field].widget.attrs['disabled'] = True

    print(realestate.addressStreet)

    context = {
        'appraisal':appraisal,
        'forms':forms,
        'realestate': realestate,
        'references': references,
        'averages': averages,
        'stds': stds,
        'tasadores':tasadores,
        'visadores':visadores,
        'appraisal_history': appraisal_history,
        'comments': comments,
        'tipoTasacion': tipoTasacion,
        'objetivo': objetivo,
        'solicitante': solicitante,
        }

    a = render(request, 'appraisal/realestate_appraisal.html', context)
    return a

def ajax_computeValuations(request):
    '''

    '''
    dict = request.GET.dict()
    UF = 30623
    liquidez = 0.9
    valorComercialUF = float(dict['valor'].strip())
    valorComercialPesos = valorComercialUF*UF
    valorLiquidezUF = valorComercialUF*liquidez
    valorLiquidezPesos = valorComercialPesos*liquidez
    montoSeguroUF = valorLiquidezUF
    montoSeguroPesos = valorLiquidezPesos
    dict = {
        "valorComercialUF": valorComercialUF,
        "valorComercialPesos": valorComercialPesos,
        "valorLiquidezUF": valorLiquidezUF,
        "valorLiquidezPesos": valorLiquidezPesos,
        "montoSeguroUF": montoSeguroUF,
        "montoSeguroPesos": montoSeguroPesos
        }
    return HttpResponse(json.dumps(dict))
