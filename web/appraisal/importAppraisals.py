from __future__ import print_function
from openpyxl import load_workbook
import re
import sys
import os
import django
#sys.path.append('/Users/Pablo Ferreiro/ProyectoInmobiliario/web/') #para pc
sys.path.append('/Users/pabloferreiro/ProyectoInmobiliario/web') #para Mac
os.environ['DJANGO_SETTINGS_MODULE'] = 'map.settings'
django.setup()

from django.core.exceptions import ObjectDoesNotExist
from django.utils.dateparse import parse_date

from realestate.models import RealEstate, Asset
from terrain.models import Terrain
from house.models import House
from building.models import Building
from apartment.models import Apartment
from appraisal.models import Appraisal, Comment, Photo, Document
from commune.models import Commune

from create.create import createOrGetRealEstate



def get_clean_address(rawaddress):
    data = {}
    address = rawaddress.strip()
    n = re.findall('\d+', address)
    regex = r'\b(#\w*[^#\W]|[^#\W]\w*#)\b'.replace('#', "D")
    d = re.split(regex, rawaddress, re.I)
    if len(d) >= 1:
        data['addressNumber2'] = d[-1].strip(". ")
        if len(d[0])>30:
            data['addressNumber2'] = d[-1].strip(". ").split('(')[-1].strip(")")
    print(n)
    if len(n) == 3:
        data['addressStreet'] = address.split(n[0])[0].strip().strip("N°n° ")
        data['addressNumber'] = n[0]
        print(data['addressNumber2'])
        print(2)
    elif len(n)== 2:
        data['addressStreet'] = address.split(n[0])[0].strip().strip("N°n° ")
        data['addressNumber'] = n[0]
        data['addressNumber2'] = n[1]

    else:
        print(len(d))
        data['addressStreet'] = address.split(n[0])[0].strip().strip("N°n° ")
        data['addressNumber'] = n[0]
        data['addressNumber2'] = None
        if len(d) > 1:
            data['addressNumber2'] = d[-1].strip(". ")
    if data['addressNumber2']==None:
        data['addressNumber2'] = ' '
    return data

def get_commune_name(commune):
    #gets commune name with first capital letter "la florida" ->"La Florida"
    name = ""
    for word in commune.split(" "):
        name += word.capitalize() + " "
    return name.strip(" ")

def green_stamp(text):
    colors = {
        'Verde':'V',
        'Amarillo':'A', 'Amarillo Vencido':'AV',
        'Rojo':'R',
        'No Aplica':'NA',
        'Verde vencido':'VV', 'Verde Vencido':'VV',
        'Sin antecedentes':'SA',
        'Sin Antecedentes':'SA',
        'No procede': 'SA', 'No encontrado':'SA'}
    return colors[text]

def excel_find_import(workbook1, workbook2, term):
    #finds data of appraisal of Santander xlxs file using santander template file
    for sheet in workbook1.sheetnames:
        for row in range(1, 400):
            for col in range(1, 100):
                cv = workbook1.get_sheet_by_name(sheet).cell(row=row, column=col).value
                if cv == term:
                    value = workbook2.get_sheet_by_name(sheet).cell(row=row, column=col).value
                    if value == None:
                        value = workbook2.get_sheet_by_name(sheet).cell(row=row+1, column=col).value
                        return value
                    else:
                        return value

def excel_find_general(file, term):
    # finds data by term in appraisal file
    print(term)
    ws = file
    if term == "PROPIEDAD ANALIZADA":
        for row in range(120, 200):
            for col in range(1, 10):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print('Found it')
                    for i in range(15):
                        terrainSquareMeters = ws.cell(row=row, column=col +15 + i).value
                        if not isinstance(terrainSquareMeters,(int, float)):
                            continue
                        else:
                            builtSquareMeters = ws.cell(row=row, column=col + 21+ i).value
                            print(terrainSquareMeters, builtSquareMeters)
                            return terrainSquareMeters, builtSquareMeters
    elif term == "VALOR COMERCIAL":
        for row in range(74, 120):
            for col in range(25, 60):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    col = 60
                    for i in range(15):
                        try:
                            valorUF = round(ws.cell(row=row, column=col-i).value,2)
                        except TypeError:
                            continue
                        if not isinstance(valorUF,(int, float)):
                            continue
                        else:
                            print(valorUF)
                            return valorUF
    elif term == "DESCRIPCIÓN GENERAL":
        for row in range(20, 50):
            for col in range(20, 50):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = " ".join(cv.split())
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    for i in range(15):
                        valor = ws.cell(row=row+1+i, column=col).value
                        if valor is not None:
                            print(valor)
                            return valor
    elif term == "DESCRIPCION SECTOR":
        for row in range(90, 120):
            for col in range(1, 50):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = " ".join(cv.split())
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    for i in range(15):
                        valor = ws.cell(row=row+4-i, column=col).value
                        if valor is not None:
                            print(valor)
                            return valor
    elif term == "Programa :":
        for row in range(100, 120):
            for col in range(1, 50):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = " ".join(cv.split())
                except AttributeError:
                    continue
                if cv == term or cv == "Programa:":
                    print("found it!")
                    for i in range(15):
                        valor = ws.cell(row=row+i+1, column=col).value
                        if valor is not None:
                            print(valor)
                            return valor
    elif term == "Estructura y Terminaciones :":
        for row in range(103, 120):
            for col in range(1, 50):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = " ".join(cv.split())
                except AttributeError:
                    continue
                if cv == term or cv == "Estructura y Terminaciones:":
                    print("found it!")
                    for i in range(15):
                        valor = ws.cell(row=row+i+1, column=col).value
                        if valor is not None:
                            print(valor)
                            return valor
    elif term == "Antigüedad" or term == "Vida Util" or term == "Sello de Gases" or term == "Tipo Propiedad":
        for row in range(35, 70):
            for col in range(1, 30):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    col = 6
                    feature = ws.cell(row=row, column=col).value
                    return feature
    elif term == "Total Avalúo Fiscal" or term == "N° Rol Principal" or term== "N° Rol (es) Sec.":
        for row in range(35, 70):
            for col in range(10, 30):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    col = 17
                    avaluo = ws.cell(row=row, column=col).value
                    return avaluo
    elif term == "Leyes que se Acoge":
        print(term)
        for row in range(35, 70):
            for col in range(5, 20):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    col = 10
                    ley1 = ws.cell(row=row, column=col).value
                    ley2 = ws.cell(row=row + 1, column=col).value
                    return ley1, ley2
    elif term == "Permiso Edificación":
        print(term)
        for row in range(35, 70):
            for col in range(5, 20):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    col = 10
                    permiso = ws.cell(row=row, column=col).value
                    fecha  = ws.cell(row=row, column=col + 1).value
                    return permiso, fecha
    elif term == "II. DESCRIPCIÓN GENERAL DEL BIEN TASADO":
        print(term)
        for row in range(14, 20):
            for col in range(1, 10):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    descripcion = ws.cell(row=row + 1, column=col).value
                    return descripcion
    elif term == "Sub Total Terreno" or term == "Sub Total Construcciones":
        print(term)
        for row in range(40, 70):
            for col in range(8, 15):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    metros = ws.cell(row=row, column=col - 4).value
                    return metros
    elif term == "Valor Comercial":
        print(term)
        for row in range(85, 100):
            for col in range(11, 20):
                cv = ws.cell(row=row, column=col).value
                try:
                    cv = cv.strip()
                except AttributeError:
                    continue
                if cv == term:
                    print("found it!")
                    valorUf = ws.cell(row=row, column=col + 3).value
                    liquidacion = ws.cell(row=row + 1, column=col + 2).value
                    return valorUf, liquidacion
    elif term == "RECEPCION FINAL N°" or term == "EXPROPIACION" or term == "VIVIENDA SOCIAL" or term == "CONST. DE ADOBE" \
            or term == "CONST. DESMONTABLES":
        if term == "RECEPCION FINAL N°":
            for row in range(20, 50):
                for col in range(30, 50):
                    cv = ws.cell(row=row, column=col).value
                    try:
                        cv = " ".join(cv.split())
                    except AttributeError:
                        continue
                    if cv == term or cv == "R. FINAL N°":
                        print("found it!")
                        valor = ws.cell(row=row, column=col + 9).value
                        return valor
        else:
            for row in range(20, 50):
                for col in range(30, 50):
                    cv = ws.cell(row=row, column=col).value
                    try:
                        cv = " ".join(cv.split())
                    except AttributeError:
                        continue
                    if cv == term:
                        print("found it!")
                        valor = ws.cell(row=row, column=col + 9).value
                        return valor

def findFromDescription(text):
    baños = 0
    dormitorios = 0
    counter = True
    numbers = {'un':1, 'uno':1, 'dos':2, 'tres':3, 'cuatro':4,'cinco':5,
               'seis':6, 'siete':7, 'ocho':8, 'nueve':9, 'dies':10}
    for i in range(len(text.split(' '))):
        word = text.split(' ')[i].lower().strip(',').strip('.').strip(' ')
        counter = True
        if word == 'baño' or word == 'baños':
            try:
                baños += int(text.split(' ')[i-1].lower().strip(',').strip('.'))
                counter = False
                continue
            except ValueError:
                if text.split(' ')[i-1].lower().strip(',').strip('.') in numbers.keys():
                    baños += numbers[text.split(' ')[i-1].lower().strip(',').strip('.')]
                    counter = False
                    continue
            try:
                baños += int(text.split(' ')[i+1])
                counter = False
                continue
            except ValueError:
                if text.split(' ')[i+1].lower().strip(',').strip('.') in numbers.keys():
                    baños += numbers[text.split(' ')[i+1].lower().strip(',').strip('.')]
                    counter = False
                    continue
            if counter:
                baños += 1

        elif word == 'dormitorio' or word== 'dormitorios':
            print(word)
            try:
                dormitorios += int(text.split(' ')[i - 1].lower().strip(',').strip('.'))
                counter = False
                continue
            except ValueError:
                if text.split(' ')[i - 1].lower().strip(',').strip('.') in numbers.keys():
                    dormitorios += numbers[text.split(' ')[i - 1].lower().strip(',').strip('.')]
                    counter = False
                    continue
            try:
                dormitorios += int(text.split(' ')[i + 1])
                counter = False
                continue
            except ValueError:
                if text.split(' ')[i + 1].lower().strip(',').strip('.') in numbers.keys():
                    dormitorios += numbers[text.split(' ')[i + 1].lower().strip(',').strip('.')]
                    counter = False
                    continue
            if counter:
                dormitorios += 1

    print(baños, dormitorios)
    return baños, dormitorios

def importAppraisalSantander(file):

    def convert(tude):
        #convert degrees, minutes, secods coordiantes into decimals

        try:
            multiplier = 1 if tude[-1] in ['N', 'E'] else -1
            d = float(tude[:-1].split('°')[0].replace(',','.'))
            m = float(tude[:-1].split('°')[1].split("'")[0].replace(',','.'))/60
            s =tude[:-1].split('°')[1].split("'")[1].split("''")[0].replace(',','.')
            s = re.sub('[^\d\.]', '', s)
            s = float(s)/3600
            print((d + m + s)*multiplier)
            return (d + m + s)*multiplier
        except ValueError:
            return 0.0

    def tipoPropiedad(text):
        tipos ={
            'Casa': Building.TYPE_CASA,
            'Pc. Agrado - Eriazo': Building.TYPE_PARCELA,
            'Departamento': Building.TYPE_DEPARTAMENTO,
            'Pc. Agrorresidencial': Building.TYPE_TERRENO,
        }
        return tipos[text]

    def ocupantes_choices(ocupante):
        #converts from file format to database format
        ocupante = ocupante.capitalize()
        if ocupante=='Propietario':
            ocupanteFinal='P'
        elif ocupante=='Arrendatario':
            ocupanteFinal='A'
        elif ocupante=='Sin ocupante':
            ocupanteFinal='SO'
        else:
            ocupanteFinal='O'
        return ocupanteFinal

    def destinoSII_modified(destino):
        # converts from file format to database format
        listaSII= {'H-Habitacional':'H','HABITACIONAL':'H',
                'O-Oficina':'O', 'OFICINA':'O',
                'C-Comercio':'C', 'COMERCIO':'C', 'COMERCIAL':'C',
                'I-Industria':'I', 'INDUSTRIA':'I',
                'L-Bodega':'L', 'BODEGA':'L',
                'Z-Estacionamiento':'Z', 'ESTACIONAMIENTO':'Z',
                'D-Deportes y Recreación':'D','DEPORTES Y RECREACIÓN':'D',
                'E-Educación y Cultura':'E','EDUCACIÓN Y CULTURA':'E',
                'G-Hotel, Motel':'G','HOTEL, MOTEL':'G',
                'P-Administración pública':'P','ADMINISTRACIÓN PÚBLICA':'P',
                'Q-Culto':'Q', 'CULTO':'Q',
                'S-Salud':'S', 'SALUD':'S',
                'SITIO ERIAZO': 'SO',
                'PROFESIONAL (oficina)':'O'   }
        return(listaSII[destino])

    def date_to_datetimefield(rawdate):
        try:
            finalDate = rawdate.split('//')[-1].strip(' ').split('-')[-1] + '-' + \
                                 rawdate.split('//')[-1].strip(' ').split('-')[-2] + \
                                 '-' + rawdate.split('//')[-1].strip(' ').split('-')[-3]
        except IndexError:
            finalDate = "1" + "-" + "1" + "-" +rawdate.split(' ')[-1]
        return parse_date(finalDate)

    def boolean_null_choices(choice):
        choice = choice.strip()
        options={
            "S/A":1, "S/Ant.":1,
            "Si":2, "SI":2,
            "No":3, "NO":3,
            "No Aplica":1, " ":1, "":1}
        return options[choice]

    def law_to_database(text):
        law ={
            'O.G.U. y C.':0,
            'P.R.C.':1,
            'Ley Pereira':2,
            'Ley 19583': 3,
            'Ley 19667':4,
            'Ley 19727':5,
            'Ley 20251':6,
            'Ley 6071':7, 'Ley N° 6071':7,
            'Ninguna':8,
            'Antigüedad':9
        }
        return law[text]

    module_dir = os.path.dirname(__file__)  # get current directory
    file_path = os.path.join(module_dir, 'static/appraisal/santander-template.xlsx')
    wb = load_workbook(filename=file_path)
    wb2 = load_workbook(filename=file, read_only=True, data_only=True)
    ws = wb2.worksheets[0]
    solicitanteCodigo = excel_find_import(wb, wb2, "solicitanteCodigo")
    id = excel_find_import(wb, wb2, "id")
    timeModified = excel_find_import(wb, wb2, "timeModified")
    solicitanteSucursal = excel_find_import(wb, wb2, "solicitanteSucursal")
    solicitanteEjecutivo = excel_find_import(wb, wb2, "solicitanteEjecutivo")
    rol = excel_find_import(wb, wb2, "rol")
    cliente = excel_find_import(wb, wb2, "cliente")
    clienteRut = excel_find_import(wb, wb2, "clienteRut")
    propietario = excel_find_import(wb, wb2, "propietario")
    propietarioRut = excel_find_import(wb, wb2, "propietarioRut")
    address = get_clean_address(excel_find_import(wb, wb2, "address"))
    addressCommune = get_commune_name(excel_find_import(wb, wb2, "addressCommune"))
    addressRegion = excel_find_import(wb, wb2, "addressRegion")
    #tasadorUser = excel_find_import(wb, wb2, "tasadorUser")
    #tasadorUserrut = excel_find_import(wb, wb2, "tasadorUser.rut")
    lat = convert(excel_find_import(wb, wb2, "lat"))
    lng = convert(excel_find_import(wb, wb2, "lng"))
    mercadoObjetivo = boolean_null_choices(excel_find_import(wb, wb2, "mercadoObjetivo"))
    antiguedad = excel_find_import(wb, wb2, "antiguedad")
    vidaUtil = excel_find_import(wb, wb2, "vidaUtil")
    if vidaUtil != int or float:
        vidaUtil = 0
    avaluoFiscal = excel_find_import(wb, wb2, "avaluoFiscal")
    acogidaLey = law_to_database(excel_find_import(wb, wb2, "acogidaLey"))
    dfl2 = boolean_null_choices(excel_find_import(wb, wb2, "dfl2"))
    selloVerde = green_stamp(excel_find_import(wb, wb2, "selloVerde"))
    copropiedadInmobiliaria = boolean_null_choices(excel_find_import(wb, wb2, "copropiedadInmobiliaria"))
    ocupante = ocupantes_choices(excel_find_import(wb, wb2, "ocupante"))
    tipoBien = excel_find_import(wb, wb2, "tipoBien")
    destinoSII = destinoSII_modified(excel_find_import(wb, wb2, "destinoSII"))
    usoActual = destinoSII_modified(excel_find_import(wb, wb2, "usoActual"))
    usoFuturo = destinoSII_modified(excel_find_import(wb, wb2, "usoFuturo"))
    permisoEdificacion = excel_find_import(wb, wb2, "permisoEdificacion")
    recepcionFinal = date_to_datetimefield(excel_find_general(ws, "RECEPCION FINAL N°"))
    expropiacion = boolean_null_choices(excel_find_general(ws, "EXPROPIACION"))
    viviendaSocial = boolean_null_choices(excel_find_general(ws, "VIVIENDA SOCIAL"))
    adobe = boolean_null_choices(excel_find_general(ws, "CONST. DE ADOBE"))
    desmontable = boolean_null_choices(excel_find_general(ws, "CONST. DESMONTABLES"))
    generalDescription = excel_find_general(ws, "DESCRIPCIÓN GENERAL")
    descripcionSectorAll = excel_find_general(ws, "DESCRIPCION SECTOR")
    programa = excel_find_general(ws, "Programa :")
    #estructuraTerminaciones = excel_find_general(file, "Estructura y Terminaciones :")
    mm2 = excel_find_general(ws, "PROPIEDAD ANALIZADA")
    terrainSquareMeters = mm2[0]
    usefulSquareMeters = mm2[0]
    builtSquareMeters = mm2[1]
    terraceSquareMeters = mm2[1]
    valorUF = excel_find_general(ws, "VALOR COMERCIAL")
    habitaciones = findFromDescription(programa)
    banos = habitaciones[0]
    dormitorios = habitaciones[1]
    #hardcoded for now
    ws2 = wb2.worksheets[0]
    propertyType = tipoPropiedad(ws2['U5'].value)

    # Crear Realestate

    propiedad = createOrGetRealEstate(
        addressNumber=address['addressNumber'],
        addressStreet=address['addressStreet'],
        addressCommune=Commune.objects.get(name=addressCommune),
        addressRegion=Commune.objects.get(name=addressCommune).region)
    propiedad.lat = lat
    propiedad.lng = lng
    print(propiedad)



    if propertyType == Building.TYPE_CASA:
        casa = propiedad.createOrGetCasa(addressNumber2=address['addressNumber2'])
        casa.bedrooms = dormitorios
        casa.bathrooms = banos
        casa.builtSquareMeters = builtSquareMeters
        casa.terrainSquareMeters = terrainSquareMeters
        casa.generalDescription = generalDescription
        casa.marketPrice = valorUF
        casa.building.name = str(file)
        # casa.building.marketPrice = valorUF
        casa.building.vidaUtilRemanente = vidaUtil
        casa.building.dfl2 = dfl2
        casa.building.avaluoFiscal = avaluoFiscal
        casa.building.copropiedadInmobiliaria = copropiedadInmobiliaria
        casa.building.selloVerde = selloVerde
        casa.building.permisoEdificacionNo = permisoEdificacion
        casa.building.permisoEdificacionFecha = recepcionFinal
        casa.building.tipoBien = tipoBien
        casa.building.rol = rol
        casa.building.year = recepcionFinal
        casa.building.mercadoObjetivo = mercadoObjetivo
        casa.building.antiguedad = antiguedad
        casa.building.acogidaLey = acogidaLey
        casa.building.ocupante = ocupante
        casa.building.adobe = adobe
        casa.building.expropiacion = expropiacion
        casa.building.viviandaSocial = viviendaSocial
        casa.building.desmontable = desmontable
        casa.building.usoActual = usoActual
        casa.building.usoFuturo = usoFuturo
        casa.building.destinoSII = destinoSII

        casa.save()
    elif propertyType == Building.TYPE_DEPARTAMENTO:
        departamento = propiedad.createOrGetDepartamento(addressNumber2=address['addressNumber2'])
        # departamento.floor = floor
        # departamento.orientation = orientation
        departamento.bedrooms = dormitorios
        departamento.bathrooms = banos
        departamento.usefulSquaremeters = usefulSquareMeters
        departamento.terraceSquaremeters = terraceSquareMeters
        departamento.generalDescription = generalDescription
        departamento.marketPrice = valorUF
        departamento.programa = programa
        departamento.apartment_building.name = str(file)
        # departamento.apartment_building.marketPrice = valorUF
        departamento.apartment_building.vidaUtilRemanente = vidaUtil
        departamento.apartment_building.dfl2 = dfl2
        departamento.apartment_building.avaluoFiscal = avaluoFiscal
        departamento.apartment_building.copropiedadInmobiliaria = copropiedadInmobiliaria
        departamento.apartment_building.selloVerde = selloVerde
        departamento.apartment_building.permisoEdificacionNo = permisoEdificacion
        departamento.apartment_building.permisoEdificacionFecha = recepcionFinal
        departamento.apartment_building.tipoBien = tipoBien
        departamento.apartment_building.rol = rol
        departamento.apartment_building.year = recepcionFinal
        departamento.apartment_building.mercadoObjetivo = mercadoObjetivo
        departamento.apartment_building.antiguedad = antiguedad
        departamento.apartment_building.acogidaLey = acogidaLey
        departamento.apartment_building.ocupante = ocupante
        departamento.apartment_building.adobe = adobe
        departamento.apartment_building.expropiacion = expropiacion
        departamento.apartment_building.viviandaSocial = viviendaSocial
        departamento.apartment_building.desmontable = desmontable
        departamento.apartment_building.usoActual = usoActual
        departamento.apartment_building.usoFuturo = usoFuturo
        departamento.apartment_building.destinoSII = destinoSII

        departamento.save()

    elif propertyType == Building.TYPE_TERRENO:
        terreno = Terrain(name=str(file), area=terrainSquareMeters, rol=rol)
        terreno.marketPrice = valorUF
        terreno.save()
        propiedad.terrains = terreno
        propiedad.save()

    print(propertyType)

    # crear tasación

    try:
        appraisal = Appraisal.objects.get(real_estates=propiedad,
                                          tipoTasacion=1,
                                          state=0,
                                          source=1,
                                          solicitanteCodigo=solicitanteCodigo,
                                          timeFinished=timeModified
                                          )
        appraisal.solicitante = 2
        appraisal.visita = 1
        appraisal.solicitanteOtro = id
        appraisal.solicitanteEjecutivo = solicitanteEjecutivo
        appraisal.solicitanteSucursal = solicitanteSucursal
        appraisal.cliente = cliente
        appraisal.clienteRut = clienteRut
        appraisal.propietario = propietario
        appraisal.propietarioRut = propietarioRut
        appraisal.descripcionSector = descripcionSectorAll
        #appraisal.liquidez = valorLiquidez
        #appraisal.tasadorUser = tasadorUser
        appraisal.timeFinished = timeModified
        appraisal.save()
        print('Existe')

    except ObjectDoesNotExist:
        appraisal = Appraisal(state=0,
                              source=1,
                              tipoTasacion=1,
                              solicitante=2,
                              visita=1,
                              solicitanteCodigo=solicitanteCodigo,
                              solicitanteOtro=id,
                              timeFinished=timeModified,
                              solicitanteEjecutivo=solicitanteEjecutivo,
                              solicitanteSucursal=solicitanteSucursal,
                              cliente=cliente,
                              clienteRut=clienteRut,
                              propietario=propietario,
                              propietarioRut=propietarioRut,
                              descripcionSector=descripcionSectorAll,
                              #liquidez=valorLiquidez
                              #tasadorUser=tasadorUser,
                              )

        appraisal.save()
        appraisal.real_estates.add(propiedad)
        appraisal.save()
        print('No Existe')

def importAppraisalITAU(file):

    def tipoPropiedad(text):
        tipos ={
            'Casa': Building.TYPE_CASA,
            'Casa en Parcela de Agrado': Building.TYPE_PARCELA,
            'Departamento': Building.TYPE_DEPARTAMENTO,
            'Terreno Unifamiliar': Building.TYPE_TERRENO,
        }
        return tipos[text]

    def law_to_database(text, text2, law):
        if law == "DFL2" and text == "DFL. 2 /59 (Viv. Económica)" or text2 == "DFL. 2 /59 (Viv. Económica)":
            return 2
        elif law == 'copropiedad' and text == "L. 19537 /97 (Cop. Inmobiliaria)" or text2 == "L. 19537 /97 (Cop. Inmobiliaria)":
            return 2
        else:
            return 0

    def estadoPropiedad(text):
        estado = {
            'Nueva':0,
            'Usada':1,
            'No Aplica':3,
            'Sitio Eriazo':3
        }
        return estado[text]


    module_dir = os.path.dirname(__file__)  # get current directory
    file_path = os.path.join(module_dir, 'static/appraisal/itau-template.xlsx')
    wb = load_workbook(filename=file_path)
    wb2 = load_workbook(filename=file, read_only=True, data_only=True)
    ws = wb2.worksheets[0]

    solicitanteCodigo = excel_find_import(wb, wb2, "solicitanteCodigo")
    id = excel_find_import(wb, wb2, "id")
    solicitanteEjecutivo = excel_find_import(wb, wb2, "solicitanteEjecutivo")
    cliente = excel_find_import(wb, wb2, "cliente")
    clienteRut = str(excel_find_import(wb, wb2, "clienteRut")) + '-' +  str(excel_find_import(wb, wb2, "num1"))
    propietario = excel_find_import(wb, wb2, "propietario")
    propietarioRut = str(excel_find_import(wb, wb2, "propietarioRut")) + '-' +  str(excel_find_import(wb, wb2, "num1"))
    addressStreet = excel_find_import(wb, wb2, "addressStreet")
    addressNumber = excel_find_import(wb, wb2, "addressNumber")
    addressNumber2 = excel_find_import(wb, wb2, "addressNumber2")
    addressCommune = get_commune_name(excel_find_import(wb, wb2, "addressCommune"))
    addressRegion = excel_find_import(wb, wb2, "addressRegion")
    rol1 = excel_find_general(file, "N° Rol Principal")
    fechaVisita = excel_find_import(wb, wb2, "timeModified")
    print(fechaVisita)
    rol2 = excel_find_general(ws, "N° Rol (es) Sec.")
    #tasadorUser = excel_find_import(wb, wb2, "tasadorUser")
    # lat = convert(excel_find_import(wb, wb2, "lat")) #Itau no viene con lat-long, usar función?
    # lng = convert(excel_find_import(wb, wb2, "lng"))
    antiguedad = int(excel_find_general(ws, "Antigüedad"))
    vidaUtil = int(excel_find_general(ws, "Vida Util"))
    vidaUtilRemanente = vidaUtil-antiguedad
    avaluoFiscal = excel_find_general(ws, "Total Avalúo Fiscal")
    leyes = excel_find_general(ws, "Leyes que se Acoge")
    acogidaLey = leyes[0]
    acogidaLey2 = leyes[1]
    selloVerde = green_stamp(excel_find_general(ws, "Sello de Gases"))
    tipoBien = estadoPropiedad(excel_find_general(ws, "Tipo Propiedad"))
    permiso = excel_find_general(ws, "Permiso Edificación")
    permisoEdificacion = permiso[0]
    recepcionFinal = permiso[1]
    generalDescription = excel_find_general(ws, "II. DESCRIPCIÓN GENERAL DEL BIEN TASADO")
    terrainSquareMeters = excel_find_general(ws, "Sub Total Terreno")
    builtSquareMeters = excel_find_general(ws, "Sub Total Construcciones")
    propertyType = tipoPropiedad(excel_find_import(wb, wb2, "propertyType"))
    valores = excel_find_general(ws, "Valor Comercial")
    valorUF = valores[0]
    valorLiquidez = valores[1]
    dfl2 = law_to_database(acogidaLey, acogidaLey2, "DFL2")
    copropiedadInmobiliaria = law_to_database(acogidaLey, acogidaLey2, "copropiedad")


    #Crear Realestate


    propiedad = createOrGetRealEstate(
        addressNumber=addressNumber,
        addressStreet=addressStreet,
        addressCommune=Commune.objects.get(name=addressCommune),
        addressRegion=Commune.objects.get(name=addressCommune).region)
    print(propiedad)

    if propertyType == Building.TYPE_CASA:
        casa = propiedad.createOrGetCasa(addressNumber2=addressNumber2)
        #casa.bedrooms = bedrooms
        #casa.bathrooms = bathrooms
        casa.builtSquareMeters = builtSquareMeters
        casa.terrainSquareMeters = terrainSquareMeters
        casa.generalDescription = generalDescription
        casa.marketPrice = valorUF
        casa.building.name = str(file)
        #casa.building.marketPrice = valorUF
        casa.building.vidaUtilRemanente = vidaUtilRemanente
        casa.building.dfl2 = dfl2
        casa.building.avaluoFiscal = avaluoFiscal
        casa.building.copropiedadInmobiliaria = copropiedadInmobiliaria
        casa.building.selloVerde = selloVerde
        casa.building.permisoEdificacionNo = permisoEdificacion
        casa.building.permisoEdificacionFecha = recepcionFinal
        casa.building.tipoPropiedad = tipoBien
        casa.building.rol = rol1
        casa.building.year = recepcionFinal

        casa.save()
    elif propertyType == Building.TYPE_DEPARTAMENTO:
        departamento = propiedad.createOrGetDepartamento(addressNumber2=addressNumber2)
        #departamento.floor = floor
        #departamento.orientation = orientation
        #departamento.bedrooms = bedrooms
        #departamento.bathrooms = bathdrooms
        departamento.usefulSquaremeters = builtSquareMeters
        departamento.generalDescription = generalDescription
        departamento.marketPrice = valorUF
        departamento.apartment_building.name = str(file)
        # departamento.apartment_building.marketPrice = valorUF
        departamento.apartment_building.vidaUtilRemanente = vidaUtilRemanente
        departamento.apartment_building.dfl2 = dfl2
        departamento.apartment_building.avaluoFiscal = avaluoFiscal
        departamento.apartment_building.copropiedadInmobiliaria = copropiedadInmobiliaria
        departamento.apartment_building.selloVerde = selloVerde
        departamento.apartment_building.permisoEdificacionNo = permisoEdificacion
        departamento.apartment_building.permisoEdificacionFecha = recepcionFinal
        departamento.apartment_building.tipoPropiedad = tipoBien
        departamento.apartment_building.rol = rol1
        departamento.apartment_building.year = recepcionFinal

        departamento.save()

    elif propertyType == Building.TYPE_TERRENO:
        terreno = Terrain(name=str(file), area=terrainSquareMeters, rol=rol1)
        terreno.marketPrice = valorUF
        terreno.save()
        propiedad.terrains=terreno
        propiedad.save()

    print(propertyType)

    # crear tasación

    try:
        appraisal = Appraisal.objects.get(real_estates=propiedad,
                                          tipoTasacion=1,
                                          state=0,
                                          source=1,
                                          solicitanteCodigo=solicitanteCodigo,
                                          timeFinished=fechaVisita
                                          )
        appraisal.solicitante = 3
        appraisal.solicitanteOtro = id
        appraisal.solicitanteEjecutivo = solicitanteEjecutivo
        appraisal.cliente = cliente
        appraisal.clienteRut = clienteRut
        appraisal.propietario = propietario
        appraisal.propietarioRut = propietarioRut
        appraisal.visita = 1
        #appraisal.liquidez = valorLiquidez
        # appraisal.tasadorUser=tasadorUser
        appraisal.save()
        print('Existe')

    except ObjectDoesNotExist:
        appraisal = Appraisal(state=0,
                              source=1,
                              tipoTasacion=1,
                              solicitante=3,
                              visita=1,
                              solicitanteCodigo=solicitanteCodigo,
                              solicitanteOtro=id,
                              timeFinished=fechaVisita,
                              solicitanteEjecutivo=solicitanteEjecutivo,
                              cliente=cliente,
                              clienteRut=clienteRut,
                              propietario=propietario,
                              propietarioRut=propietarioRut,
                              #liquidez=valorLiquidez
                              #tasadorUser=tasadorUser,
                              )

        appraisal.save()
        appraisal.real_estates.add(propiedad)
        appraisal.save()
        print('No Existe')



files_santander =  ['N-1775585 (15930247-4) Av. La Florida 9650 Casa 60 Altos de Santa Amalia La Florida inc min promesa 19-10-18.xlsx',
'N-1777974 (16336209-0) Santa Isabel 797 dp 1016 Santiago.xlsx',
'N-1774960 (13267388-8) Hernán Olguín 0359, Los Heroes Poniente Maipú.xlsx',
'N-1775293 (9314398-1) Consistorial 2608 (Via Amarilla) casa 14, Condominio Casas del Consistorial Peñalolen.xlsx',
'N-1775307 (13450369-6) Transit 480 Block 2-B Estacion Central.xlsx',
'N-1775460 (12516411-0) Las Violetas 2152  Dpto 407 Providencia Rol 2428-25 (E 33) (A 2017) (7P).xlsx',
'N-1775763 (15421021-0) Luis Pereira 1621 Dpto E Ñuñoa Rol 1851-88 (T 26) (E 77) (A 1995).xlsx',
'N-1777517 (12874278-6) Río Teno 1069 Villa Bahia Catalina La Granja Rol 5935-8 (T 86 E 57) (A 2009).xlsx',
'N-1777660 (16713130-1) Credito 596 Providencia.xlsx',
'N-1777834 (21254788-3) Tarapacá 782, Dp 206, Santiago.xlsx',
'N-1775967 (77557450-K) Lo Lopez 1469 Cerro Navia Rol 62851 (T 816) Terreno.xlsx']
files_itau = []
file_mac = '/Volumes/GoogleDrive/Mi unidad/ProyectoInmobiliario/Datos/tasaciones/'
file_pc = 'G:/Mi unidad/ProyectoInmobiliario/Datos/tasaciones/'

for dir in files_santander:
    file = file_mac + dir
    print(file)
    importAppraisalSantander(file)

#importAppraisalITAU(file)
'''

path = file_pc

files = os.listdir(path)
for name in files:
    print(name)
'''

