from django.views.generic import FormView
from django.shortcuts import render
from django.core import serializers
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.utils.text import slugify
from django.contrib.auth.decorators import login_required
from django.core.exceptions import MultipleObjectsReturned

from .forms import AppraisalCreateForm, GrupoCreateForm

from region.models import Region
from commune.models import Commune
from condominium.models import Condominium
from square.models import Square
from realestate.models import RealEstate
from building.models import Building
from appraisal.models import Appraisal, Rol
from . import create
from . import parse

from openpyxl import drawing
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook

import datetime

import requests

import fitz
import zipfile

from bs4 import BeautifulSoup

from lxml import html

@login_required(login_url='/user/login')
def view_create(request):

    if request.method == 'POST':

        request_post = request.POST.copy()
        request_post['appraisalTimeRequest'] = datetime.datetime.strptime(request_post['appraisalTimeRequest'],'%d/%m/%Y %H:%M')
        request_post['appraisalTimeDue'] = datetime.datetime.strptime(request_post['appraisalTimeDue'],'%d/%m/%Y %H:%M')
        
        form = AppraisalCreateForm(request_post)
        if form.is_valid():

            propertyType = int(form.cleaned_data['propertyType'])

            # 1. Crear real estate
            real_estate, created = create.createOrGetRealEstate(
                addressNumber=form.cleaned_data['addressNumber'],
                addressStreet=form.cleaned_data['addressStreet'],
                addressCommune=form.cleaned_data['addressCommune'],
                addressRegion=form.cleaned_data['addressRegion'])

            #   Political divisions of land
            real_estate.addressLoteo = form.cleaned_data['addressLoteo']
            real_estate.addressSitio = form.cleaned_data['addressSitio']
            if form.cleaned_data['addressSquare'] != "" and form.cleaned_data['addressSquare'] != None:
                try:
                    square = Square.objects.get(code=form.cleaned_data['addressSquare'])
                except Condominium.DoesNotExist:
                    square = Square(code=form.cleaned_data['addressSquare'])
                    square.save()
                real_estate.addressSquare = square

            #   Cultural identifiers
            for grupo in range(sum(1 for key in request_post.keys() if key.startswith("addressCondominiumType"))):
                if request_post["addressCondominiumText_{}".format(grupo+1)] != "":
                    addressCondominiumType = request_post["addressCondominiumType_{}".format(grupo+1)]
                    addressCondominiumText = request_post["addressCondominiumText_{}".format(grupo+1)]
                    real_estate.addressCondominium.create(name=addressCondominiumText,ctype=addressCondominiumType)

            real_estate.save()

            # 2. Crear apraisal
            if request.FILES:
                orderFile = request.FILES['archivo']
            else:
                orderFile = None
            appraisal = create.createAppraisal(request,real_estate,
                solicitante=form.cleaned_data['solicitante'],
                solicitanteOtro=form.cleaned_data['solicitanteOther'],
                solicitanteSucursal=form.cleaned_data['solicitanteSucursal'],
                solicitanteCodigo=form.cleaned_data['solicitanteCodigo'],
                solicitanteEjecutivo=form.cleaned_data['solicitanteEjecutivo'],
                solicitanteEjecutivoEmail=form.cleaned_data['solicitanteEjecutivoEmail'],
                solicitanteEjecutivoTelefono=form.cleaned_data['solicitanteEjecutivoTelefono'],
                timeDue=form.cleaned_data['appraisalTimeDue'],
                timeRequest=form.cleaned_data['appraisalTimeRequest'],
                tipoTasacion=form.cleaned_data['tipoTasacion'],
                finalidad=form.cleaned_data['finalidad'],
                visita=form.cleaned_data['visita'],
                cliente=form.cleaned_data['cliente'],
                clienteRut=form.cleaned_data['clienteRut'],
                clienteEmail=form.cleaned_data['clienteEmail'],
                clienteTelefono=form.cleaned_data['clienteTelefono'],
                contacto=form.cleaned_data['contacto'],
                contactoEmail=form.cleaned_data['contactoEmail'],
                contactoTelefono=form.cleaned_data['contactoTelefono'],
                price=None,
                commentsOrder=form.cleaned_data['comments'],
                orderFile=orderFile)

            # 3. Agregar terrenos y construcciones
            
            # crear el rol
            if len(form.cleaned_data['rol']) > 0:
                rol = Rol(code=form.cleaned_data['rol'])
            else:
                rol = Rol()

            # crear propiedad y asignar rol
            propertyType = int(form.cleaned_data['propertyType'])
            appraisal.property_main_type = propertyType
            if propertyType == Building.TYPE_TERRENO:
                propiedad, created = real_estate.createOrGetTerreno(addressNumber2=form.cleaned_data['addressNumber2'])
                appprop = appraisal.addAppProperty(Building.TYPE_TERRENO,propiedad.id)
                appraisal.property_main = appprop
                rol.terrain = propiedad
            elif propertyType == Building.TYPE_CASA:
                propiedad, created = real_estate.createOrGetTerreno(addressNumber2=form.cleaned_data['addressNumber2'])
                appraisal.addAppProperty(Building.TYPE_TERRENO,propiedad.id) # en general se tasa el terreno Y la casa
                propiedad, created = real_estate.createOrGetCasa(addressNumber2=form.cleaned_data['addressNumber2'])
                appprop = appraisal.addAppProperty(Building.TYPE_CASA,propiedad.id)
                appraisal.property_main = appprop
                rol.house = propiedad
            elif propertyType == Building.TYPE_EDIFICIO:
                propiedad, created = real_estate.createOrGetTerreno(addressNumber2=form.cleaned_data['addressNumber2'])
                appraisal.addAppProperty(Building.TYPE_TERRENO,propiedad.id)
                propiedad, created = real_estate.createOrGetEdificio(addressNumber2=form.cleaned_data['addressNumber2'])
                appprop = appraisal.addAppProperty(Building.TYPE_EDIFICIO,propiedad.id)
                appraisal.property_main = appprop
                rol.apartment_building = propiedad
            elif propertyType == Building.TYPE_DEPARTAMENTO:
                propiedad, created = real_estate.createOrGetDepartamento(addressNumber2=None,addressNumber3=form.cleaned_data['addressNumber2'])
                appprop = appraisal.addAppProperty(Building.TYPE_DEPARTAMENTO,propiedad.id)
                appraisal.property_main = appprop
                rol.apartment = propiedad
            elif propertyType == Building.TYPE_LOCAL_COMERCIAL:
                propiedad, created = real_estate.createOrGetTerreno(addressNumber2=form.cleaned_data['addressNumber2'])
                appraisal.addAppProperty(Building.TYPE_TERRENO,propiedad.id)
                propiedad, created = real_estate.createOrGetLocalComercial(addressNumber2=form.cleaned_data['addressNumber2'])
                appprop = appraisal.addAppProperty(Building.TYPE_LOCAL_COMERCIAL,propiedad.id)
                appraisal.property_main = appprop
                rol.local_comercial = propiedad
            else:
                """
                Otro tipo de propiedades de las que no tenemos clases.
                """
                pass
            
            rol.save()
            appraisal.save()
            
            # Go to appraisal url
            #return HttpResponseRedirect(appraisal.url)
            # Go to appraisal main
            return HttpResponseRedirect('/list/')

        else:
            errordata = form.errors.as_data()
            if '__all__' in errordata.keys():
                message = errordata['__all__'][0].message
            else:
                message = ""
            context = {'form':form,'message':message}            
            return render(request, 'create/error.html', context)

    else:

        # Form
        form = AppraisalCreateForm(label_suffix='')
        communes = Commune.objects.only('name').order_by('name')
        regions = Region.objects.only('name').order_by('code')
        form.fields['addressRegion'].queryset = regions
        form.fields['addressCommune'].queryset = communes

        context = {'form': form}

        return render(request, 'create/index.html', context)

def load_communes(request):
    region_id = int(request.GET.get('region'))
    region = Region.objects.get(pk=region_id)
    communes = list(Commune.objects.filter(region=region.code).order_by('name'))
    return render(request,
        'hr/commune_dropdown_list_options.html',
        {'communes': communes})

def import_request(request):

    data = {}

    url = request.POST['url']

    if 'archivo' not in request.FILES.keys() and url == "":

        data['error'] = 'Debe elegir un archivo o ingresar una url antes de importar.'
        return JsonResponse(data)

    if url != "":

        data = parse.parseSantanderUrl(url)

    if 'archivo' in request.FILES.keys():

        file = request.FILES['archivo']
        filetype = file._name.split('.')[1]

        if filetype == 'xls':

            data['error'] = "No es posible importar archivos excel '.xls'. Se recomienda guardar el archivo en formato '.xlsx'."

        elif filetype == 'xlsx':

            # Itau o Chile Avances

            try:
                wb = load_workbook(filename=file,read_only=True,data_only=True)
            except zipfile.BadZipFile:
                data['error'] = "Archivo parece estar asegurado. Se recomienda abrir y volver a guardar el archivo."
                return JsonResponse(data)            
            ws = wb.worksheets[0]
            if ws['C1'].value != None and 'SOLICITUD DE TASACIÓN' in ws['C1'].value.strip():
                data = parse.parseItau(ws)
            elif ws['B5'].value != None and 'SOLICITUD DE INFORME' in ws['B5'].value.strip():
                data = parse.parseBancoDeChileAvance(wb)
            else:
                data['error'] = "Formato de tasación no reconocido."

        elif filetype == 'pdf':

            # Santander o Chile

            file = file.read()
            doc = fitz.open(filetype="pdf",stream=file)
            page = doc.loadPage(0)
            text = page.getText("text").splitlines()
            if 'Solicito a usted' in text[0]:
                data = parse.parseBancoDeChile(text)
            elif 'Nº Req' in text[0]:
                data = parse.parseSantander(text)
            else:
                data['error'] = "Formato de tasación no reconocido."

        else:

            data['error'] = "Formato de archivo no reconocido."

    return JsonResponse(data)

def ajax_load_grupo(request):
    form = GrupoCreateForm
    return render(request,'create/grupo.html',{'form': form})