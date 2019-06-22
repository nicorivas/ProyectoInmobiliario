#!/usr/bin/env python
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select



from openpyxl import drawing
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook

import time

def read_excel(wb):

    ws = wb.get_sheet_by_name("TASACIÓN")

    excel_to_variable_dict = {
        "O14": "solicitanteSucursal",
        "O15": "solicitanteEjecutivo",
        "O16": "cliente",
        "O17": "clienteRut",
        "O18": "propietario",
        "O19": "propietarioRut",
        "O20": "addressStreet",
        "O21": "addressCondominium",
        "O22": "rol",
        "O23": "addressComuna",
        "AA23": "addressCiudad",
        "O24": "addressRegion",
        "O25": "tasador",
        "O26": "tasadorRut",
        "O27": "tasadorEmpresa",
        "AB35": "descripcionGeneral",
        "AY14": "mercadoObjetivo",
        "AY15": "antiguedad",
        "AY16": "vidaUtilRemanente",
        "AY17": "avaluoFiscal",
        "AY18": "acogidaA",
        "AY19": "DFL2",
        "AY20": "selloVerde",
        "AY21": "copopriedadInmobiliaria",
        "AY22": "ocupante",
        "AY23": "destinoSegunSII",
        "AY24": "tipoBien",
        "AY25": "usoActual",
        "AY26": "usoFuturo",
        "AY27": "permisoEdificacion",
        "AY28": "recepcionFinal",
        "AY29": "expropiacion",
        "AY30": "construccionDesmontables",
        "AY31": "adobe",
        "AY87": "rentaMensual",
        "AY88": "rentaMeses",
        "AY90": "rentaTasa",
        "B87": "rentaDescripcion",
        "W177": "porcentajeLiquidez",
        "I223": "zonaEdificacion",
        "I224": "zonaUsos",
        "I225": "usosPermitidos",
        "I231": "usosProhibidos",
        "AK223": "superficiePredialMinima",
        "AK224": "frentePredialMinimo",
        "AK225": "coeficienteOcupacionSuelo",
        "AK226": "coeficienteMaximoConstructibilidad",
        "AK227": "alturaMaxima",
        "AK228": "estacionamientosCirculacion",
        "AK229": "densidadBrutaMaxima",
        "AK230": "antejardin",
        "AK231": "distanciamientoMinimoAMedianeros",
        "AK232": "sistemaAgrupamiento",
        "AK233": "distanciaCierroEdificacion"
    }
    app = {}
    for key, value in excel_to_variable_dict.items():
        print(key,value,ws[key].value)
        app[value] = ws[key].value

    # Terrenos
    terrenos = []
    for i, c in enumerate(range(99,104)):
        if ws["B"+str(c)].value == None:
            break
        terrenos.append(
            {
            "nombre":ws["B"+str(c)].value,
            "unidad":ws["U"+str(c)].value,
            "superficie":ws["W"+str(c)].value,
            "ufm2":ws["AF"+str(c)].value
            })
    app["terrenos"] = terrenos
    
    # Construcciones
    construcciones = []
    for i, c in enumerate(range(106,114)):
        if ws["B"+str(c)].value == None:
            break
        construcciones.append(
            {
            "nombre":ws["B"+str(c)].value,
            "material":ws["K"+str(c)].value,
            "ano":ws["N"+str(c)].value,
            "prenda":ws["P"+str(c)].value,
            "recepcion":ws["R"+str(c)].value,
            "unidad":ws["U"+str(c)].value,
            "superficie":ws["W"+str(c)].value
            })
    app["construcciones"] = construcciones

    # Construcciones
    obras = []
    for i, c in enumerate(range(116,120)):
        if ws["B"+str(c)].value == None:
            break
        obras.append(
            {
            "nombre":ws["B"+str(c)].value,
            "valorUF":ws["AR"+str(c)].value
            })
    app["obras"] = obras

    return app

def fill_web_inner(dictionary):
    '''
    '''    

def fill_web(dictionary):

    global browser
    
    # === Init

    browser = webdriver.Chrome()

    # Login
    browser.get("https://extranet.gruposantander.cl/")
    el = browser.find_element_by_id("close")
    el.click()
    user = browser.find_element_by_name("user")
    password = browser.find_element_by_name("clave")
    user.send_keys("70149761")
    password.send_keys("protasa2018")
    submit = browser.find_element_by_id("btnSbmt")
    submit.click()

    # Main
    browser.get("https://extranet.gruposantander.cl/gateW.aspx?dst_url=https://tasaciones.extranetsantander.cl/segesta/login_prov_adm.aspx?qry=ok&UrlToken=&Target=principal&UsaToken=SI")
    browser.switch_to.frame("iframe2")

    # Tasaciones en visacion
    browser.find_element_by_partial_link_text('Por enviar a Visado').click()

    # Seleccionar tasacion (agarrar el popup)
    main_window_handle = None
    while not main_window_handle:
        main_window_handle = browser.current_window_handle
    appraisal_link = browser.find_element_by_partial_link_text('3651')
    appraisal_link.click()
    appraisal_window_handle = None
    while not appraisal_window_handle:
        for handle in browser.window_handles:
            if handle != main_window_handle:
                appraisal_window_handle = handle
                break
    browser.switch_to.window(appraisal_window_handle)
    
    # === Identificacion

    browser.find_element_by_partial_link_text('IDENTIFICA').click()
    
    if dictionary["tasadorRut"] != None:
        browser.find_element_by_id("rut_tasador").send_keys(dictionary["tasadorRut"].replace(".",""))
    
    if dictionary["rol"] != None:
        browser.find_element_by_xpath('//span[text()="Agregar Nuevo Rol"]').click()
        browser.find_element_by_id("txt_rolIzquierda").send_keys(dictionary["rol"].split("-")[0])
        browser.find_element_by_id("txt_rolDerecha").send_keys(dictionary["rol"].split("-")[1])
        browser.find_element_by_xpath('//button[text()="Agregar"]').click()
        time.sleep(1)
        browser.find_element_by_xpath('//h2[text()="Rol ingresado exitosamente."]/../div[contains(@class,"sa-button-container")]//button[text()="OK"]').click()
        Select(browser.find_element_by_id("ddl_Roles")).select_by_value("{:05d}-{:05d}".format(int(dictionary["rol"].split("-")[0]),int(dictionary["rol"].split("-")[1])))

    if dictionary["descripcionGeneral"] != None:
        browser.find_element_by_name("txtDescripcionGeneral").send_keys(dictionary["descripcionGeneral"])

    if dictionary["expropiacion"] != None:
        Select(browser.find_element_by_id("ddl_Expropiacion")).select_by_value(dictionary["expropiacion"].title())
    
    if dictionary["mercadoObjetivo"] != None:
        Select(browser.find_element_by_id("ddl_mercado")).select_by_value(dictionary["mercadoObjetivo"].title())

    if dictionary["propietarioRut"] != None:
        browser.find_element_by_name("txt_RutPropietario").send_keys(dictionary["propietarioRut"])
    
    if dictionary["propietario"] != None:
        browser.find_element_by_name("txt_NomPropietario").send_keys(dictionary["propietario"])

    browser.find_element_by_id("btn_Guardar").click()

    browser.find_element_by_xpath('//h2[text()="Información guardada de forma exitosa."]/../div[contains(@class,"sa-button-container")]//button[text()="OK"]').click()
    
    time.sleep(1)

    # === Valorizacion

    browser.find_element_by_partial_link_text('VALORIZA').click()

    def add_terenos():
        for terreno in dictionary["terrenos"]:
            time.sleep(1)
            browser.find_element_by_id("addTerr").click()
            time.sleep(1)
            browser.find_element_by_id("txt_descterreno").send_keys(terreno["nombre"])
            browser.find_element_by_id("txt_supterreno").send_keys(terreno["superficie"])
            browser.find_element_by_xpath('//button[text()="Guardar Terreno"]').click()
            time.sleep(1)
            browser.find_element_by_xpath('//h2[text()="Terreno ingresado exitosamente."]/../div[contains(@class,"sa-button-container")]//button[text()="Cerrar"]').click()
    
    def add_construcciones():
        for construccion in dictionary["construcciones"]:
            time.sleep(1)
            browser.find_element_by_id("addEdif").click()
            time.sleep(1)
            if construccion["nombre"] != None:
                browser.find_element_by_id("txt_detalleitem").send_keys(construccion["nombre"])
                browser.find_element_by_id("txt_descripcionedif").send_keys(construccion["nombre"])
            if construccion["superficie"] != None:
                browser.find_element_by_id("txt_edifSuperficie").send_keys(construccion["superficie"])
            if construccion["ano"] != None:
                browser.find_element_by_id("txt_anoCons").send_keys(construccion["ano"])
            Select(browser.find_element_by_id("ddl_Prenda")).select_by_value('1')
            if construccion["material"] == "A-Acero":
                Select(browser.find_element_by_id("ddl_materialidad")).select_by_visible_text("A - Acero")
            elif construccion["material"] == "B-Hormigon":
                Select(browser.find_element_by_id("ddl_materialidad")).select_by_visible_text("B - Horm. Armado")
            elif construccion["material"] == "C-Albañileria":
                Select(browser.find_element_by_id("ddl_materialidad")).select_by_visible_text("C - Alb. Ladrillo/Bloque")
            elif construccion["material"] == "D-Piedra/Bloque":
                Select(browser.find_element_by_id("ddl_materialidad")).select_by_visible_text("D - Albañileria Piedra")
            elif construccion["material"] == "E-Madera":
                Select(browser.find_element_by_id("ddl_materialidad")).select_by_visible_text("E - Madera")
            elif construccion["material"] == "F-Adobe":
                Select(browser.find_element_by_id("ddl_materialidad")).select_by_visible_text("F - Adobe")
            elif construccion["material"] == "G-Metalcom":
                Select(browser.find_element_by_id("ddl_materialidad")).select_by_visible_text("G - Pref. Acero")
            elif construccion["material"] == "H-Prefab. Madera":
                Select(browser.find_element_by_id("ddl_materialidad")).select_by_visible_text("H - Pref. Madera")
            elif construccion["material"] == "I-Prefab. Hormigón":
                Select(browser.find_element_by_id("ddl_materialidad")).select_by_visible_text("I - Pref. Hormigon")
            elif construccion["material"] == "J-Otros":
                Select(browser.find_element_by_id("ddl_materialidad")).select_by_visible_text("J - Otras")
            browser.find_element_by_xpath('//button[text()="Guardar Edificación"]').click()
            time.sleep(1)
            browser.find_element_by_xpath('//h2[text()="Edificación ingresada exitosamente."]/../div[contains(@class,"sa-button-container")]//button[text()="Cerrar"]').click()
    
    def add_obras():
        for obra in dictionary["obras"]:
            time.sleep(1)
            browser.find_element_by_id("addOOCC").click()
            time.sleep(1)
            if obra["nombre"] != None:
                browser.find_element_by_id("txt_descoocc").send_keys(obra["nombre"])
            browser.find_element_by_xpath('//button[text()="Guardar OOCC"]').click()
            time.sleep(1)
            browser.find_element_by_xpath('//h2[text()="OOCC ingresada exitosamente."]/../div[contains(@class,"sa-button-container")]//button[text()="Cerrar"]').click()

    add_terenos()
    add_construcciones()
    add_obras()

    Select(browser.find_element_by_id("ddl_porcLiq")).select_by_value(str(int(dictionary["porcentajeLiquidez"]*100)))

    # Renta
    if dictionary["rentaMensual"] != None:
        browser.find_element_by_id("txt_Renta").send_keys(dictionary["rentaMensual"])
    if dictionary["rentaMeses"] != None:
        Select(browser.find_element_by_id("ddl_ArrMeses")).select_by_value(dictionary["rentaMeses"])
    if dictionary["rentaTasa"] != None:
        Select(browser.find_element_by_id("ddl_ArrTasa")).select_by_value(dictionary["rentaTasa"])
    if dictionary["rentaDescripcion"] != None:
        browser.find_element_by_id("txt_ComentarioRenta").send_keys(dictionary["rentaDescripcion"])
    browser.find_element_by_id("btnGuardaRenta").click()
    time.sleep(1)
    browser.find_element_by_xpath('//h2[text()="Observación ingresada exitosamente."]/../div[contains(@class,"sa-button-container")]//button[text()="Cerrar"]').click()

    time.sleep(1)

    # === Sector

    browser.find_element_by_partial_link_text('SECTOR').click()

    if dictionary["zonaEdificacion"] != None:
        browser.find_element_by_id("txtCodigoArea").send_keys(dictionary["zonaEdificacion"])
    if dictionary["zonaUsos"] != None:
        browser.find_element_by_id("txtTipoArea").send_keys(dictionary["zonaUsos"])
    if dictionary["coeficienteMaximoConstructibilidad"] != None:
        browser.find_element_by_id("txtCoeficiente").send_keys(dictionary["coeficienteMaximoConstructibilidad"])
    if dictionary["sistemaAgrupamiento"] != None:
        Select(browser.find_element_by_id("ddlSistemaAgrupamiento")).select_by_value(dictionary["sistemaAgrupamiento"])
    if dictionary["coeficienteOcupacionSuelo"] != None:
        browser.find_element_by_id("txtOcupacionSuelo").send_keys(dictionary["coeficienteOcupacionSuelo"])
    if dictionary["alturaMaxima"] != None:
        browser.find_element_by_id("txtAlturaEdificacion").send_keys(dictionary["alturaMaxima"])
    if dictionary["densidadBrutaMaxima"] != None:
        browser.find_element_by_id("txtDensidad").send_keys(dictionary["densidadBrutaMaxima"])